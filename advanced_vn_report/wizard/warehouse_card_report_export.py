from odoo import fields, models, api
import base64
import logging
from copy import copy
from io import BytesIO
from datetime import datetime

from odoo import fields, models, _, api
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource

from openpyxl.cell.cell import MergedCell
from openpyxl.styles.borders import Border

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')

class WarehouseCardReportFile (models.TransientModel):
    _name = 'warehouse.card.report.file'
    file_name = fields.Char()
    file = fields.Binary()

class WarehouseCardReportExport (models.TransientModel):
    _name = 'warehouse.card.report.export'
    _description = 'Description'

    def _get_default_location(self):
        locations = self.env['stock.location'].search([])
        list_location = []
        if len(locations) > 0:
            for location in locations:
                if location:
                    if location.usage == 'internal':
                        location_fake_ids = self.env['stock.location'].sudo().search(
                            [('parent_path', '=like', location.parent_path + '%')])
                        if len(location_fake_ids) == 1:
                            for location_fake in location_fake_ids:
                                if location_fake:
                                    list_location.append(location_fake.id)
        if len(list_location) > 0:
            return [(6, 0, list_location)]
        else:
            return False

    location_children_ids = fields.Many2many('stock.location', '', string='Cac dia diem con',
                                             default=_get_default_location)
    location_id = fields.Many2one('stock.location', string='Nhà kho')
    product_ids = fields.Many2many('product.product', string='Sản phẩm')
    date_from = fields.Date('Từ ngày')
    date_to = fields.Date('Đến ngày')

    @api.constrains('date_to', 'date_from')
    def _check_date_to_from(self):
        for rec in self:
            if rec.date_to < rec.date_from:
                raise ValidationError(_('Ngày bắt đầu phải trước ngày kết thúc.'))

    def confirm_export(self):
        wb = load_workbook(get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                               'Thẻ kho.xlsx'))
        ws = wb['Sheet1']
        index = 1
        if len(self.product_ids.ids) > 1:
            index_sheet = 2
            so_sheet_them = len(self.product_ids.ids) - 1
            for i in range(so_sheet_them):
                new_sheet = wb.copy_worksheet(ws)
                new_sheet.title = 'Sheet' + str(index_sheet)
                index_sheet += 1
        for product in self.product_ids:
            self.fill_data(workbook=wb, worksheet=wb['Sheet' + str(index)], product=product)
            index += 1
        content = BytesIO()
        wb.save(content)
        out = base64.encodestring(content.getvalue())
        view = self.env.ref('advanced_vn_report.warehouse_card_report_file_form_view')
        self.env['warehouse.card.report.file'].sudo().search([]).unlink()
        file_output = self.env['warehouse.card.report.file'].sudo().create({
            'file': out,
            'file_name': "Thẻ kho.xlsx"
        })
        content.close()
        return {
            'name': "Thẻ kho",
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'warehouse.card.report.file',
            'target': 'new',
            'view_id': view.id,
            'res_id': file_output.id
        }

    def fill_data(self, workbook, worksheet, product):
        global_style = copy(worksheet['A6']._style)
        worksheet['A5'] = "Ngày lập thẻ: " + datetime.today().strftime("%d/%m/%Y")
        worksheet['E7'] = product.name
        worksheet['E8'] = product.uom_id.name if product.uom_id else None
        worksheet[
            'E9'] = product.product_tmpl_id.default_code if product.product_tmpl_id and product.product_tmpl_id.default_code else None
        worksheet['A5']._style = global_style
        worksheet['E7']._style = global_style
        worksheet['E8']._style = global_style
        worksheet['E9']._style = global_style
        sml = self.env['stock.move.line'].sudo().search(
            [('product_id', '=', product.id), '|', ('location_dest_id', '=', self.location_id.id),
             ('location_id', '=', self.location_id.id), ('state', '=', 'done'), ('date', '>=', self.date_from),
             ('date', '<=', self.date_to)])
        insert_rows = len(sml)
        while insert_rows > 0:
            worksheet.insert_rows(13)
            worksheet.cell(row=13, column=1)._style = copy(worksheet.cell(row=14, column=1)._style)
            worksheet.cell(row=13, column=2)._style = copy(worksheet.cell(row=14, column=2)._style)
            worksheet.cell(row=13, column=3)._style = copy(worksheet.cell(row=14, column=3)._style)
            worksheet.cell(row=13, column=4)._style = copy(worksheet.cell(row=14, column=4)._style)
            worksheet.cell(row=13, column=5)._style = copy(worksheet.cell(row=14, column=5)._style)
            worksheet.cell(row=13, column=6)._style = copy(worksheet.cell(row=14, column=6)._style)
            worksheet.cell(row=13, column=7)._style = copy(worksheet.cell(row=14, column=7)._style)
            worksheet.cell(row=13, column=8)._style = copy(worksheet.cell(row=14, column=8)._style)
            worksheet.cell(row=13, column=9)._style = copy(worksheet.cell(row=14, column=9)._style)
            worksheet.cell(row=13, column=10)._style = copy(worksheet.cell(row=14, column=10)._style)
            insert_rows -= 1
        # global_style_2 = copy(worksheet['B14']._style)
        # product.invalidate_cache()
        # so ton dau ky
        ton = product.with_context(location=self.location_id.id,
                                               to_date=self.date_from).qty_available
        nhap = 0
        xuat = 0
        index = 1
        for line in sml:
            if line.location_id.id == self.location_id.id:
                ton -= line.qty_done
                xuat += line.qty_done
                worksheet.cell(row=13+index-1, column=1).value = index
                worksheet.cell(row=13+index-1, column=2).value = line.move_id.date.strftime("%d/%m/%Y")
                worksheet.cell(row=13+index-1, column=6).value = line.move_id.date.strftime("%d/%m/%Y")
                worksheet.cell(row=13+index-1, column=4).value = line.reference
                worksheet.cell(row=13+index-1, column=5).value = line.move_id.name
                worksheet.cell(row=13+index-1, column=8).value = line.qty_done
                worksheet.cell(row=13+index-1, column=9).value = ton
                index += 1
            if line.location_dest_id.id == self.location_id.id:
                ton += line.qty_done
                nhap += line.qty_done
                worksheet.cell(row=13+index-1, column=1).value = index
                worksheet.cell(row=13+index-1, column=2).value = line.move_id.date.strftime("%d/%m/%Y")
                worksheet.cell(row=13+index-1, column=6).value = line.move_id.date.strftime("%d/%m/%Y")
                worksheet.cell(row=13+index-1, column=3).value = line.reference
                worksheet.cell(row=13+index-1, column=5).value = line.move_id.name
                worksheet.cell(row=13+index-1, column=7).value = line.qty_done
                worksheet.cell(row=13+index-1, column=9).value = ton
                index += 1
        worksheet.cell(row=13 + index - 1, column=5).value = 'Cộng cuối kỳ'
        worksheet.cell(row=13 + index - 1, column=7).value = nhap
        worksheet.cell(row=13 + index - 1, column=8).value = xuat
        worksheet.cell(row=13 + index - 1, column=9).value = ton