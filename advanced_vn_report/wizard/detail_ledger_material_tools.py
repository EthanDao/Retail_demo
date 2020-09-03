import base64
import logging
from copy import copy
from io import BytesIO
from datetime import datetime

from odoo import fields, models, _, api
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource

from openpyxl.cell.cell import MergedCell
from openpyxl.styles.borders import Border

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class DetailLedgerMaterialTools(models.Model):
    _name = 'detail.ledger.material.tools'

    date_from = fields.Date('Từ ngày')
    date_to = fields.Date('Đến ngày')

    # in_year = fields.Date(string="Năm")
    def _get_default_location(self):
        locations = self.env['stock.location'].search([])
        list_location = []
        if len(locations) > 0:
            for location in locations:
                if location:
                    if location.usage == 'internal':
                        location_fake_ids = self.env['stock.location'].sudo().search(
                            [('parent_path', '=like', location.parent_path + '%')])
                        if len(location_fake_ids) == 1:
                            for location_fake in location_fake_ids:
                                if location_fake:
                                    list_location.append(location_fake.id)
        if len(list_location) > 0:
            return [(6, 0, list_location)]
        else:
            return False

    location_children_ids = fields.Many2many('stock.location', 'location_children_ids_rel', string='Cac dia diem con',
                                             default=_get_default_location)
    location_ids = fields.Many2many('stock.location', string='Nhà kho')
    product_ids = fields.Many2many('product.product', string='Sản phẩm')
    unit_id = fields.Many2one(comodel_name="uom.uom", string="Đơn vị tính")

    @api.constrains('date_to', 'date_from')
    def _check_date_to_from(self):
        for rec in self:
            if rec.date_to < rec.date_from:
                raise ValidationError(_('Ngày bắt đầu phải trước ngày kết thúc.'))

    def confirm_export_detail_ledger(self):
        wb = load_workbook(get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                               'Sổ chi tiết vật liệu.xlsx'))
        ws = wb['Sheet1']
        index = 1
        if len(self.location_ids.ids) > 1:
            index_sheet = 2
            so_sheet_them = len(self.location_ids.ids) - 1
            for i in range(so_sheet_them):
                new_sheet = wb.copy_worksheet(ws)
                new_sheet.title = 'Sheet' + str(index_sheet)
                index_sheet += 1
        for location in self.location_ids:
            self.fill_data(workbook=wb, worksheet=wb['Sheet' + str(index)], location=location)
            index += 1
        content = BytesIO()
        wb.save(content)
        out = base64.encodestring(content.getvalue())
        view = self.env.ref('advanced_vn_report.result_detail_ledger_material_tools_form_view')
        self.env['result.detail.ledger.material.tools'].sudo().search([]).unlink()
        file_output = self.env['result.detail.ledger.material.tools'].sudo().create({
            'file': out,
            'file_name': "Sổ chi tiết vật liệu.xlsx"
        })
        content.close()
        return {
            'name': "Sổ chi tiết vật liệu",
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'result.detail.ledger.material.tools',
            'target': 'new',
            'view_id': view.id,
            'res_id': file_output.id
        }

    def fill_data(self, workbook, worksheet, location):
        worksheet['A6'] = "Từ ngày " + str(self.date_from.strftime("%d/%m/%Y")) + " đến ngày " + str(
            self.date_to.strftime("%d/%m/%Y"))
        worksheet['A7'] = "Mã kho: " + str(location.location_id.display_name) + "            " + "Tên kho: " + str(
            location.name)
        stock_valuation = self.env['stock.valuation.layer'].sudo().search(
            [('create_date', '>=', self.date_from), ('create_date', '<=', self.date_to)])
        # count_move_line =
        worksheet['J8'] = "Đơn vị tính: " + str(stock_valuation.currency_id.name if stock_valuation.currency_id else '')
        row = 12
        col = 1
        index = 1
        global_style = copy(worksheet['A12']._style)
        global_style_1 = copy(worksheet['A13']._style)
        global_style_2 = copy(worksheet['B14']._style)
        global_style_3 = copy(worksheet['C14']._style)
        global_style_4 = copy(worksheet['E14']._style)
        # tim sml co nhap hoac xuat
        sml_start = self.env['stock.move.line'].sudo().search(
            ['|', ('location_dest_id', '=', location.id), ('location_id', '=', location.id)])
        # dem so product kem them dong ton dau ky
        # count_product = 0
        # lay ra danh sach san pham co nhap va xuat
        # product_list = []
        # for sml in sml_start:
        #     if sml.product_id.type == 'product':
        #         if sml.product_id in self.product_ids:
        #             product_list.append(sml.product_id)
        #             count_product += 2
        # count_move_line = 0
        # lay so move line trong svl
        list_move_line = []
        # so luong thay doi gia trong product
        list_price_new = []
        count_list_price_new = 0
        # so luong thay doi categ product
        list_change_product_category = []
        count_change_product_category = 0
        if stock_valuation:
            for valuation in stock_valuation:
                if valuation.product_id in self.product_ids:
                    if valuation.is_change_standard_price:
                        list_price_new.append(valuation.id)
                        count_list_price_new += 2
                    if valuation.is_change_product_category:
                        if valuation.account_move_id:
                            list_change_product_category.append(valuation.id)
                            count_change_product_category += 2
                    if valuation.stock_move_id.move_line_ids:
                        for line in valuation.stock_move_id.move_line_ids:
                            if line.location_id.id == location.id or line.location_dest_id.id == location.id:
                                list_move_line.append(line)
        sum_insert_row = len(self.product_ids) * 2 + count_list_price_new + len(
            list_move_line) + count_change_product_category
        if sum_insert_row >= 8:
            so_dong_them = sum_insert_row - 8
            start_insert_row = 19
            while so_dong_them > 0:
                worksheet.insert_rows(start_insert_row)
                for coltmp in range(1, 13):
                    if coltmp in range(2, 11, 2):
                        worksheet.cell(row=start_insert_row, column=coltmp)._style = global_style_2
                    elif coltmp in [1, 3, 12]:
                        worksheet.cell(row=start_insert_row, column=coltmp)._style = global_style_3
                    elif coltmp in range(5, 12, 2):
                        worksheet.cell(row=start_insert_row, column=coltmp)._style = global_style_4
                start_insert_row = start_insert_row + 1
                so_dong_them -= 1
        total_in_end = 0
        total_out_end = 0
        # total_money_available_end = 0
        for product in self.product_ids:
            # tong ton thanh tien
            total_money_available_end = 0
            # money_in = 0
            # money_out = 0

            # xoa cache truoc khi chuyen location
            product.invalidate_cache()
            # so ton dau ky
            quant_available = product.with_context(location=location.id,
                                                   to_date=self.date_from).qty_available
            # # thanh tien nhap
            # money_in = 0
            # # thanh tien xuat
            # money_out = 0
            # don gia sp thay doi
            change_price = 0
            # tong nhap
            sum_in_end = 0
            # tong xuat
            sum_out_end = 0
            # so ton cuoi ky
            quant_available_end = 0
            worksheet.cell(row=row, column=col).value = index
            index += 1
            # ten san pham
            worksheet.cell(row=row, column=col).value = "Mã hàng: " + str(
                product.default_code if product.default_code else " ") + "                                    Tên hàng: " + str(
                product.name)
            worksheet.cell(row=row, column=col)._style = global_style_1
            row += 1
            worksheet.cell(row=row, column=col).value = "- Số dư đầu kỳ"
            # lay gia cu cua product
            svl_old_price = self.env['stock.valuation.layer'].search(
                [('old_price', '!=', False), ('old_price', '>', 0), ('create_date', '>=', self.date_from),
                 ('create_date', '<=', self.date_to),
                 ('product_id', '=', product.id)], limit=1)
            thanh_tien_dau_ky = 0
            if svl_old_price:
                # don gia
                worksheet.cell(row=row, column=col + 4).value = '{:,.2f}'.format(svl_old_price.old_price)
                # so ton dau ky
                worksheet.cell(row=row, column=col + 9).value = quant_available
                # so ton dau ky thanh tien
                thanh_tien_dau_ky = quant_available * svl_old_price.old_price
                worksheet.cell(row=row, column=col + 10).value = '{:,.2f}'.format(thanh_tien_dau_ky)
            else:
                # don gia
                worksheet.cell(row=row, column=col + 4).value = '{:,.2f}'.format(product.standard_price)
                # so ton dau ky
                worksheet.cell(row=row, column=col + 9).value = quant_available
                # so ton dau ky thanh tien
                thanh_tien_dau_ky = quant_available * product.standard_price
                worksheet.cell(row=row, column=col + 10).value = '{:,.2f}'.format(thanh_tien_dau_ky)

            if stock_valuation:

                for valuation in stock_valuation:
                    # thanh tien nhap
                    money_in = 0
                    # thanh tien xuat
                    money_out = 0
                    if valuation.stock_move_id:
                        for line in valuation.stock_move_id.move_line_ids:
                            if len(line) > 0:
                                if line.location_id.id == location.id or line.location_dest_id.id == location.id:
                                    if valuation.product_id == product:
                                        # worksheet.cell(row=row, column=col)._style = global_style_2
                                        row += 1
                                        if valuation.account_move_id:
                                            # so hieu
                                            worksheet.cell(row=row, column=col).value = valuation.account_move_id.name
                                            # dien giai
                                            worksheet.cell(row=row,
                                                           column=col + 2).value = valuation.account_move_id.ref
                                            # date
                                            worksheet.cell(row=row,
                                                           column=col + 1).value = valuation.account_move_id.date.strftime(
                                                "%d/%m/%Y")
                                        # don gia
                                        if valuation.quantity != 0:
                                            cost = valuation.value / valuation.quantity
                                            # don gia
                                            if cost == 0:
                                                worksheet.cell(row=row,
                                                               column=col + 4).value = '0.00'
                                            else:
                                                worksheet.cell(row=row,
                                                               column=col + 4).value = '{:,.2f}'.format(cost)
                                        # so luong nhap
                                        if line.location_dest_id.id == location.id:
                                            # tai khoan doi ung
                                            if valuation.stock_move_id:
                                                if valuation.stock_move_id.reference:
                                                    # Nhap kho NVL thua, lay related account bt
                                                    if 'NKNVLT' in valuation.stock_move_id.reference:
                                                        for x in valuation.account_move_id.vn_line_ids:
                                                            if len(x) > 0:
                                                                if x.debit > 0:
                                                                    # account.append(x)
                                                                    worksheet.cell(row=row,
                                                                                   column=col + 3).value = x.related_account_id.code if len(
                                                                        x) > 0 else ''
                                                    # Nhap KNVL thanh pham lay account input trong product categ
                                                    else:
                                                        worksheet.cell(row=row,
                                                                       column=col + 3).value = product.categ_id.property_stock_account_input_categ_id.code if product.categ_id else ''
                                            worksheet.cell(row=row, column=col + 5).value = line.qty_done
                                            if valuation.quantity != 0:
                                                # don gia
                                                cost = valuation.value / valuation.quantity
                                                # nhap thanh tien
                                                if cost == 0:
                                                    worksheet.cell(row=row,
                                                                   column=col + 6).value = '0.00'
                                                else:
                                                    worksheet.cell(row=row,
                                                                   column=col + 6).value = '{:,.2f}'.format(
                                                        line.qty_done * cost)
                                                # tong nhap thanh tien
                                                money_in = line.qty_done * cost
                                                total_in_end += line.qty_done * valuation.value / valuation.quantity
                                                sum_in_end += line.qty_done

                                            # else:
                                            #     worksheet.cell(row=row, column=col + 5).value = line.qty_done
                                            #     # nhap thanh tien
                                            #     worksheet.cell(row=row,
                                            #                    column=col + 6).value = '{:,.0f}₫'.format(
                                            #         valuation.value)
                                        # so luong xuat
                                        elif line.location_id.id == location.id:
                                            # tai khoan doi ung
                                            for x in valuation.account_move_id.vn_line_ids:
                                                if len(x) > 0:
                                                    if x.credit > 0:
                                                        # account.append(x)
                                                        worksheet.cell(row=row,
                                                                       column=col + 3).value = x.related_account_id.code if len(
                                                            x) > 0 else ''
                                            if valuation.quantity != 0:
                                                # so luong xuat
                                                worksheet.cell(row=row, column=col + 7).value = line.qty_done
                                                # don gia
                                                cost = valuation.value / valuation.quantity
                                                # xuat thanh tien
                                                if cost == 0:
                                                    worksheet.cell(row=row,
                                                                   column=col + 8).value = '0.00'
                                                else:
                                                    worksheet.cell(row=row,
                                                                   column=col + 8).value = '{:,.2f}'.format(
                                                        line.qty_done * cost)
                                                # tong xuat thanh tien
                                                money_out = abs(line.qty_done * cost)
                                                total_out_end += line.qty_done * valuation.value / valuation.quantity
                                                sum_out_end += line.qty_done
                                        # so ton cuoi ky = ton dau ky + sum(nhap) - sum(xuat)
                                        quant_available_end = quant_available + sum_in_end - sum_out_end
                                        worksheet.cell(row=row, column=col + 9).value = quant_available_end

                                        if valuation.quantity != 0 and change_price == 0:
                                            # tong ton thanh tien
                                            if total_money_available_end == 0:
                                                total_money_available_end = thanh_tien_dau_ky + money_in - money_out
                                            elif total_money_available_end != 0:
                                                total_money_available_end += money_in - money_out
                                            # so ton cuoi ky thanh tien
                                            worksheet.cell(row=row,
                                                           column=col + 10).value = '{:,.2f}'.format(
                                                total_money_available_end)
                                        elif valuation.quantity != 0 and change_price != 0:
                                            # tong ton thanh tien
                                            if total_money_available_end == 0:
                                                total_money_available_end = change_price + money_in - money_out
                                            elif total_money_available_end != 0:
                                                total_money_available_end += money_in - money_out
                                            # so ton cuoi ky thanh tien
                                            worksheet.cell(row=row,
                                                           column=col + 10).value = '{:,.2f}'.format(
                                                total_money_available_end)
                                        # elif valuation.current_price != 0:
                                        #     print(valuation.value / valuation.quantity)
                                        #     # tong ton thanh tien
                                        #     total_money_available_end = quant_available_end * valuation.value / valuation.quantity + money_in - money_out
                                        #     # so ton cuoi ky thanh tien
                                        #     worksheet.cell(row=row,
                                        #                    column=col + 10).value = '{:,.2f}'.format(
                                        #         total_money_available_end)
                    elif valuation.is_change_standard_price:
                        if valuation.product_id == product:
                            # khi thay doi gia ton thanh tien thay doi
                            total_money_available_end = 0
                            money_in = 0
                            money_out = 0
                            change_price = quant_available_end * valuation.current_price
                            row += 1
                            worksheet.cell(row=row, column=col)._style = global_style
                            # so tien sau khi giam or tang gia sp
                            value_price = valuation.current_price - valuation.old_price
                            worksheet.cell(row=row, column=col).value = _(
                                'Giá trị sản phẩm được sửa đổi (từ %s thành %s)') % (
                                                                            '{:,.2f}'.format(valuation.old_price),
                                                                            '{:,.2f}'.format(valuation.current_price))
                            row += 1
                            worksheet.cell(row=row, column=col)._style = global_style_2
                            # if valuation.account_move_id:
                            #     # so hieu
                            #     worksheet.cell(row=row, column=col).value = valuation.account_move_id.name
                            #     # dien giai
                            #     worksheet.cell(row=row,
                            #                    column=col + 2).value = valuation.account_move_id.ref if valuation.account_move_id.ref else ''
                            #     # date
                            #     worksheet.cell(row=row,
                            #                    column=col + 1).value = valuation.account_move_id.date.strftime(
                            #         "%d/%m/%Y")
                            #     # tai khoan doi ung
                            #     for x in valuation.account_move_id.vn_line_ids:
                            #         if len(x) > 0:
                            #             if value_price > 0:
                            #                 if x.debit > 0:
                            #                     # account.append(x)
                            #                     worksheet.cell(row=row,
                            #                                    column=col + 3).value = x.related_account_id.code if len(
                            #                         x) > 0 else ''
                            #             else:
                            #                 if x.credit > 0:
                            #                     # account.append(x)
                            #                     worksheet.cell(row=row,
                            #                                    column=col + 3).value = x.related_account_id.code if len(
                            #                         x) > 0 else ''
                            worksheet.cell(row=row, column=col + 4).value = '{:,.2f}'.format(valuation.current_price)
                            # So luong ton cuoi ky
                            worksheet.cell(row=row, column=col + 9).value = quant_available_end
                            # gia thay doi thanh tien
                            change_price = quant_available_end * valuation.current_price
                            # Thanh tien = sl cuoi ton ky * so tien sau khi giam or tang sp
                            worksheet.cell(row=row, column=col + 10).value = '{:,.2f}'.format(
                                quant_available_end * valuation.current_price)
                            # if valuation.quantity != 0:
                            #     # tong ton thanh tien
                            #     total_money_available_end = thanh_tien_dau_ky + money_in - money_out
                            #     # so ton cuoi ky thanh tien
                            #     worksheet.cell(row=row,
                            #                    column=col + 10).value = '{:,.2f}'.format(
                            #         total_money_available_end)
                    elif valuation.is_change_product_category and valuation.account_move_id:
                        if valuation.product_id == product:

                            row += 1
                            worksheet.cell(row=row, column=col)._style = global_style
                            #dien giai
                            worksheet.cell(row=row,
                                           column=col).value = valuation.description if valuation.description else ''
                            row += 1
                            worksheet.cell(row=row, column=col)._style = global_style_2
                            if valuation.account_move_id:
                                # so hieu
                                worksheet.cell(row=row, column=col).value = valuation.account_move_id.name
                                # dien giai
                                worksheet.cell(row=row,
                                               column=col + 2).value = valuation.account_move_id.ref
                                # date
                                worksheet.cell(row=row,
                                               column=col + 1).value = valuation.account_move_id.date.strftime(
                                    "%d/%m/%Y")
                            worksheet.cell(row=row, column=col + 4).value = '{:,.2f}'.format(
                                valuation.value / valuation.quantity)
                            # So luong ton cuoi ky
                            worksheet.cell(row=row, column=col + 9).value = quant_available_end
                            # gia thay doi thanh tien
                            # change_price = quant_available_end * valuation.current_price
                            # Thanh tien = sl cuoi ton ky * so tien sau khi giam or tang sp
                            worksheet.cell(row=row, column=col + 10).value = '{:,.2f}'.format(
                                quant_available_end * valuation.value / valuation.quantity)
                    worksheet.cell(row=row, column=col)._style = global_style
            row += 1
            col = 1
        if sum_insert_row >= 8:
            worksheet.cell(row=row, column=col + 6).value = '{:,.2f}'.format(total_in_end)
            worksheet.cell(row=row, column=col + 8).value = '{:,.2f}'.format(total_out_end)
            # if total_in_end > 0 or total_out_end > 0:
            #     worksheet.cell(row=row, column=col + 10).value = '{:,.2f}'.format(total_in_end - total_out_end)
            # else:
            #     worksheet.cell(row=row, column=col + 10).value = '{:,.2f}'.format(
            #         quant_available * product.standard_price)
            worksheet.merge_cells(start_row=row + 4, start_column=1, end_row=row + 4, end_column=2)
            worksheet.merge_cells(start_row=row + 5, start_column=1, end_row=row + 5, end_column=2)
            worksheet.merge_cells(start_row=row + 4, start_column=5, end_row=row + 4, end_column=6)
            worksheet.merge_cells(start_row=row + 5, start_column=5, end_row=row + 5, end_column=6)
            worksheet.merge_cells(start_row=row + 3, start_column=9, end_row=row + 3, end_column=11)
            worksheet.merge_cells(start_row=row + 4, start_column=9, end_row=row + 4, end_column=11)
            worksheet.merge_cells(start_row=row + 5, start_column=9, end_row=row + 5, end_column=11)
        else:
            worksheet.cell(row=20, column=col + 6).value = '{:,.2f}'.format(total_in_end)
            worksheet.cell(row=20, column=col + 8).value = '{:,.2f}'.format(total_out_end)
            # if total_in_end > 0 or total_out_end > 0:
            #     worksheet.cell(row=20, column=col + 10).value = '{:,.2f}'.format(total_in_end - total_out_end)
            # else:
            #     worksheet.cell(row=20, column=col + 10).value = '{:,.2f}'.format(
            #         quant_available * product.standard_price)
            worksheet.merge_cells(start_row=24, start_column=1, end_row=24, end_column=2)
            worksheet.merge_cells(start_row=25, start_column=1, end_row=25, end_column=2)
            worksheet.merge_cells(start_row=24, start_column=5, end_row=24, end_column=6)
            worksheet.merge_cells(start_row=25, start_column=5, end_row=25, end_column=6)
            worksheet.merge_cells(start_row=23, start_column=9, end_row=23, end_column=11)
            worksheet.merge_cells(start_row=24, start_column=9, end_row=24, end_column=11)
            worksheet.merge_cells(start_row=25, start_column=9, end_row=25, end_column=11)
