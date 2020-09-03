import base64
import logging
from copy import copy
from io import BytesIO
from datetime import datetime

from odoo import fields, models, _, api
from odoo.exceptions import ValidationError
from odoo.modules.module import get_module_resource
from datetime import timedelta
from openpyxl.cell.cell import MergedCell
from openpyxl.styles.borders import Border

_logger = logging.getLogger(__name__)
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class StockQuantityReportExport(models.Model):
    _name = 'stock.quantity.report.export'

    date_from = fields.Date(string='Từ ngày ')
    date_to = fields.Date(string='Đến ngày')
    location_id = fields.Many2one(comodel_name='stock.location', string='Kho', required=False)

    @api.constrains('date_to', 'date_from')
    def _check_date_to_from(self):
        for rec in self:
            if rec.date_to < rec.date_from:
                raise ValidationError(_('Ngày bắt đầu phải trước ngày kết thúc.'))

    def confirm(self):
        wb = load_workbook(get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                               'Báo cáo xuất nhập tồn.xlsx'))
        ws = wb['Sheet1']
        self.fill_data(workbook=wb, worksheet=ws)
        content = BytesIO()
        wb.save(content)
        out = base64.encodestring(content.getvalue())
        view = self.env.ref('advanced_vn_report.stock_quantity_report_file_form_view')
        self.env['stock.quantity.report.file'].sudo().search([]).unlink()
        file_output = self.env['stock.quantity.report.file'].sudo().create({
            'file': out,
            'file_name': "Báo cáo xuất nhập tồn.xlsx"
        })
        content.close()
        return {
            'name': "Báo cáo xuất nhập tồn",
            'type': 'ir.actions.act_window',
            'view_type': 'form',
            'view_mode': 'form',
            'res_model': 'stock.quantity.report.file',
            'target': 'new',
            'view_id': view.id,
            'res_id': file_output.id
        }

    def fill_data(self, workbook, worksheet):
        global_style = copy(worksheet['A6']._style)
        worksheet['B7'] = str(self.date_from.strftime("%d/%m/%Y"))
        worksheet['B8'] = str(self.date_to.strftime("%d/%m/%Y"))
        worksheet['B6'] = str(self.location_id.complete_name)
        worksheet['B6']._style = global_style
        worksheet['B7']._style = global_style
        worksheet['B8']._style = global_style
        date_from = self.date_from
        date_to = self.date_to
        location_ids = self.env['stock.location'].sudo().search([('parent_path', '=like', self.location_id.parent_path + '%' )])
        for location_id in location_ids:
            # location_id = location_id.id
            sml_start = self.env['stock.move.line'].sudo().search(
                    ['|', ('location_dest_id', '=', location_id.id),
                     ('location_id', '=', location_id.id)])
            product_list = []
            for sml in sml_start:
                if sml.product_id in product_list:
                    pass
                else:
                    if sml.product_id.type == 'product':
                        product_list.append(sml.product_id)
            for product in product_list:

                sml_qty_in = 0  # số lượng nhập trong kì
                sml_qty_out = 0  # số lượng xuất trong kì
                sml_value_in = 0  # thành tiền nhập trong kì
                sml_value_out = 0  # thành tiền nhập trong kì
                # xoa cache truoc khi truyen location
                product.invalidate_cache()
                #so luong ton dau ky
                sml_qty_start = product.with_context(location=location_id.id, to_date=date_from).qty_available
                svl_start = self.env['stock.valuation.layer'].sudo().search(
                    [('create_date', '<', date_from),
                     ('product_id', '=', product.id)])
                svl_start_qty = sum([line.quantity for line in svl_start])
                svl_start_value = sum([line.value for line in svl_start])
                svl_price_unit = 0
                if svl_start_qty != 0:
                    svl_price_unit = svl_start_value / svl_start_qty
                sml_value_start = svl_price_unit * sml_qty_start # gia thanh ton dau ky
                # # Tìm các stock move line nhập trong thời điểm lựa chọn ( nhập trong kì )
                # sml_in = self.env['stock.move.line'].sudo().search(
                #     [('create_date', '>=', date_from),
                #      ('create_date', '<=', date_to),
                #      ('location_dest_id', '=', location_id),
                #      ('product_id', '=', product.id)])
                # # Tìm các stock move line xuất trong thời điểm lựa chọn ( xuất trong kì )
                # sml_out = self.env['stock.move.line'].sudo().search(
                #     [('create_date', '>=', date_from),
                #      ('create_date', '<=', date_to),
                #      ('location_id', '=', location_id),
                #      ('product_id', '=', product.id)])
                # for sml_id in sml_in:
                #     qty = sml_id.qty_done
                #     qty_total = sum([line.qty_done for line in sml_id.move_id.move_line_ids])
                #     value_total = sum([line.amount_total_signed for line in sml_id.move_id.account_move_ids])
                #     if qty_total != 0:
                #         sml_value_in += value_total / qty_total * qty
                #     sml_qty_in += qty
                # for sml_id in sml_out:
                #     qty = sml_id.qty_done
                #     qty_total = sum([line.qty_done for line in sml_id.move_id.move_line_ids])
                #     value_total = sum([line.amount_total_signed for line in sml_id.move_id.account_move_ids])
                #     if qty_total != 0:
                #         sml_value_out += value_total / qty_total * qty
                #     sml_qty_out += qty
                # tim cac svl trong ky
                svl_list = self.env['stock.valuation.layer'].sudo().search(
                    [('create_date', '>=', self.date_from), ('create_date', '<=', self.date_to), ('product_id','=',product.id)])
                worksheet['K8'] = "Đơn vị tính: " + str(
                    product_list[0].currency_id.name if product_list[0] else '')
                for svl in svl_list:
                    if svl.stock_move_id:
                        if svl.stock_move_id.move_line_ids:
                            for move_line in svl.stock_move_id.move_line_ids:
                                if move_line.location_dest_id.id == location_id.id:
                                    qty_in = move_line.qty_done
                                    value_in = svl.unit_cost * qty_in
                                    sml_qty_in += qty_in
                                    sml_value_in += value_in
                                if move_line.location_id.id == location_id.id:
                                    qty_out = move_line.qty_done
                                    value_out = svl.unit_cost * qty_out
                                    sml_qty_out += qty_out
                                    sml_value_out += value_out

                # xoa cache truoc khi truyen location
                product.invalidate_cache()
                # tồn cuối kỳ
                sml_qty_end = product.with_context(location=location_id.id, to_date=(date_to + timedelta(days=1)) ).qty_available
                svl_end = self.env['stock.valuation.layer'].sudo().search(
                    [('create_date', '<=', date_to),
                     ('product_id', '=', product.id)])
                svl_end_qty = sum([line.quantity for line in svl_end])
                svl_end_value = sum([line.value for line in svl_end])
                svl_price_unit_end = 0
                if svl_end_qty != 0:
                    svl_price_unit_end = svl_end_value / svl_end_qty
                sml_value_end = svl_price_unit_end * sml_qty_end  # gia thanh ton cuoi ky
                if svl_start_qty == 0 and sml_qty_in == 0 and sml_qty_out == 0 and svl_end_qty == 0:
                    pass
                else:
                    # x = sml_value_end - sml_value_start - sml_value_in + sml_value_out
                    worksheet.cell(row=13, column=1).value = location_id.complete_name
                    worksheet.cell(row=13, column=1)._style = copy(worksheet.cell(row=14, column=1)._style)
                    worksheet.cell(row=13, column=2).value = product.product_tmpl_id.default_code
                    worksheet.cell(row=13, column=2)._style = copy(worksheet.cell(row=14, column=2)._style)
                    worksheet.cell(row=13, column=3).value = product.product_tmpl_id.name
                    worksheet.cell(row=13, column=3)._style = copy(worksheet.cell(row=14, column=3)._style)
                    worksheet.cell(row=13, column=4).value = product.product_tmpl_id.uom_id.name
                    worksheet.cell(row=13, column=4)._style = copy(worksheet.cell(row=14, column=4)._style)
                    worksheet.cell(row=13, column=5).value = sml_qty_start
                    worksheet.cell(row=13, column=5)._style = copy(worksheet.cell(row=14, column=5)._style)
                    worksheet.cell(row=13, column=6).value = '{:,.2f}'.format(sml_value_start)
                    worksheet.cell(row=13, column=6)._style = copy(worksheet.cell(row=14, column=6)._style)
                    worksheet.cell(row=13, column=7).value = sml_qty_in
                    worksheet.cell(row=13, column=7)._style = copy(worksheet.cell(row=14, column=7)._style)
                    worksheet.cell(row=13, column=8).value = '{:,.2f}'.format(sml_value_in)
                    worksheet.cell(row=13, column=8)._style = copy(worksheet.cell(row=14, column=8)._style)
                    worksheet.cell(row=13, column=9).value = sml_qty_out
                    worksheet.cell(row=13, column=9)._style = copy(worksheet.cell(row=14, column=9)._style)
                    worksheet.cell(row=13, column=10).value = '{:,.2f}'.format(sml_value_out)
                    worksheet.cell(row=13, column=10)._style = copy(worksheet.cell(row=14, column=10)._style)
                    # worksheet.cell(row=13, column=11).value = '{:,.2f}'.format(x)
                    # worksheet.cell(row=13, column=11)._style = copy(worksheet.cell(row=14, column=11)._style)
                    worksheet.cell(row=13, column=11).value = sml_qty_end
                    worksheet.cell(row=13, column=11)._style = copy(worksheet.cell(row=14, column=11)._style)
                    worksheet.cell(row=13, column=12).value = '{:,.2f}'.format(sml_value_end)
                    worksheet.cell(row=13, column=12)._style = copy(worksheet.cell(row=14, column=12)._style)
                    worksheet.insert_rows(13)
        worksheet.delete_rows(13)

