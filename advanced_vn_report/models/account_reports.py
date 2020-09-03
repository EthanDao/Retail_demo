# -*- coding: utf-8 -*-
# import datetime
import io
import json
import logging
from copy import copy
from datetime import datetime, timedelta

from odoo.tools import date_utils

from openpyxl.utils import get_column_letter
from dateutil.relativedelta import relativedelta
from odoo import models, _, api
from odoo.exceptions import UserError
from odoo.modules.module import get_module_resource
from odoo.tools.safe_eval import safe_eval

_logger = logging.getLogger(__name__)
import openpyxl

try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class AccountReport(models.AbstractModel):
    _inherit = 'account.report'

    def _get_reports_buttons(self):
        return [
            {'name': _('Print Preview'), 'sequence': 1, 'action': 'print_pdf', 'file_export_type': _('PDF')},
            {'name': _('Export (XLSX)'), 'sequence': 2, 'action': 'print_xlsx', 'file_export_type': _('XLSX')},
            {'name': _('Export XLSX (Theo thông tư 200)'), 'sequence': 3, 'action': 'advanced_print_xlsx',
             'file_export_type': _('XLSX')},
        ]

    def advanced_print_xlsx(self, options):
        return {
            'type': 'ir_actions_account_report_template_download',
            'data': {'model': self.env.context.get('model'),
                     'options': json.dumps(options),
                     'output_format': 'xlsx',
                     'financial_id': self.env.context.get('id'),
                     }
        }

    @api.model
    def _format_aml_name(self, line_name, move_ref, move_name):
        names = []
        if move_name != '/':
            names.append(move_name)
        if move_ref and move_ref != '/':
            names.append(move_ref)
        if line_name and line_name != '/':
            names.append(line_name)
        name = '-'.join(names)
        # TODO: check if no_format is still needed
        if len(name) > 35 and not self.env.context.get('no_format'):
            pass
        return name

    def advanced_get_xlsx(self, options, response=None, line_id=None):
        # global y_offset, new_sheet
        report_manager = self._get_report_manager(options)
        output = io.BytesIO()
        workbook = None
        now = datetime.now()
        current_year = now.year
        if self.id == self.env.ref('advanced_vn_report.account_financial_report_cash_flow_statement_b03').id:
            template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                'Báo cáo Lưu chuyển tiền tệ.xlsx')
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook['Sheet1']
            sheet['K6'] = 'Từ ngày ' + str(datetime.strptime(options['date']['date_from'], "%Y-%m-%d").date().strftime(
                '%d-%m-%Y')) + ' đến ngày ' + str(
                datetime.strptime(options['date']['date_to'], "%Y-%m-%d").date().strftime('%d-%m-%Y'))
            rows = sheet.rows
            count = 0
            data = {}
            data_1 = {}
            for row in rows:
                count += 1
                if 10 <= count <= 38:
                    for cell in row:
                        if cell.value is not None and cell.coordinate == ('P' + str(count)):
                            format = '(' + str(cell.value) + ')'
                            data.update({
                                format: ('T' + str(count))
                            })
                            data_1.update({
                                format: ('Y' + str(count))
                            })
            # todo generate data follow template
            ctx = self._set_context(options)
            ctx.update({'no_format': True, 'print_mode': True, 'prefetch_fields': False})
            lines = self.with_context(ctx)._get_lines(options)
            if len(lines) > 0:
                for line in lines:
                    if len(line['columns']) > 1:
                        if line['columns'][0]['name'] != '':
                            for key in data:
                                if key in line['name']:
                                    if line['columns'][0]['name'] >= 0:
                                        sheet[data[key]] = '{:,.0f}'.format(line['columns'][0]['name'])
                                    else:
                                        sheet[data[key]] = '(' + str(
                                            '{:,.0f}'.format(abs(line['columns'][0]['name']))) + ')'
                        if line['columns'][1]['name'] != '':
                            for key in data:
                                if key in line['name']:
                                    if line['columns'][1]['name'] >= 0:
                                        sheet[data_1[key]] = '{:,.0f}'.format(line['columns'][1]['name'])
                                    else:
                                        sheet[data_1[key]] = '(' + str(
                                            '{:,.0f}'.format(abs(line['columns'][1]['name']))) + ')'
                    else:
                        raise UserError(_('Không có dữ liệu cột so sánh theo kỳ.'))

            # print(data_1)
            # print(data)

            # dau ky
            # count = 0
            # data_1 = {}
            # rows2 = sheet.rows
            # for row2 in rows2:
            #     count += 1
            #     if 10 <= count <= 38:
            #         for cell in row2:
            #             if cell.value is not None and cell.coordinate == ('P' + str(count)):
            #                 format = '(' + str(cell.value) + ')'
            #                 data_1.update({
            #                     format: ('Y' + str(count))
            #                 })
            # date_from = datetime.strptime(options['date']['date_from'], "%Y-%m-%d")
            # date_to = datetime.strptime(options['date']['date_to'], "%Y-%m-%d")
            # period_type = options['date']['period_type']
            # if period_type == 'custom':
            #     date_dif = date_to - date_from
            #     new_date_to = date_from - timedelta(days=1)
            #     new_date_from = new_date_to - date_dif
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'quarter':
            #     new_date_from = date_from + relativedelta(months=-3)
            #     new_date_to = date_to + relativedelta(months=-3)
            #     new_date_from = date_utils.start_of(new_date_from, 'month')
            #     new_date_to = date_utils.end_of(new_date_to, 'month')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'month':
            #     new_date_from = date_from + relativedelta(months=-1)
            #     new_date_to = date_to + relativedelta(months=-1)
            #     new_date_from = date_utils.start_of(new_date_from, 'month')
            #     new_date_to = date_utils.end_of(new_date_to, 'month')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'fiscalyear':
            #     new_date_from = date_from + relativedelta(years=-1)
            #     new_date_to = date_to + relativedelta(years=-1)
            #     new_date_from = date_utils.start_of(new_date_from, 'year')
            #     new_date_to = date_utils.end_of(new_date_to, 'year')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")

            # ctx = self._set_context(options)
            # ctx.update({'no_format': True, 'print_mode': True, 'prefetch_fields': False})
            # lines2 = self.with_context(ctx)._get_lines(options)
            # for line2 in lines2:
            #     if line2['columns'][1]['name'] != '':
            #         for key in data_1:
            #             if key in line2['name']:
            #                 sheet[data_1[key]] = line2['columns'][1]['name']
        elif self.id == self.env.ref('advanced_vn_report.account_financial_report_b01').id:
            template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                'Bảng cân đối kế toán.xlsx')
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook['Sheet1']
            sheet['B7'] = 'Từ ngày ' + str(datetime.strptime(options['date']['date_from'], "%Y-%m-%d").date().strftime(
                '%d-%m-%Y')) + ' đến ngày ' + str(
                datetime.strptime(options['date']['date_to'], "%Y-%m-%d").date().strftime('%d-%m-%Y'))
            rows = sheet.rows
            count = 0
            data = {}
            data_1 = {}
            for row in rows:
                count += 1
                if 10 <= count <= 123:
                    for cell in row:
                        if cell.value is not None and cell.coordinate == ('B' + str(count)):
                            format = '(' + str(cell.value) + ')'
                            data.update({
                                format: ('D' + str(count))
                            })
                            data_1.update({
                                format: ('E' + str(count))
                            })
            # todo generate data follow template
            ctx = self._set_context(options)
            ctx.update({'no_format': True, 'print_mode': True, 'prefetch_fields': False})
            lines = self.with_context(ctx)._get_lines(options)
            if len(lines) > 0:
                for line in lines:
                    if len(line['columns']) > 1:
                        if line['columns'][0]['name'] != '':
                            for key in data:
                                if key in line['name']:
                                    sheet[data[key]] = line['columns'][0]['name']
                        if line['columns'][1]['name'] != '':
                            for key in data:
                                if key in line['name']:
                                    sheet[data_1[key]] = line['columns'][1]['name']
                    else:
                        raise UserError(_('Không có dữ liệu cột so sánh theo kỳ.'))
            # dau ky
            # count = 0
            # data_1 = {}
            # rows2 = sheet.rows
            # for row2 in rows2:
            #     count += 1
            #     if 10 <= count <= 123:
            #         for cell in row2:
            #             if cell.value is not None and cell.coordinate == ('B' + str(count)):
            #                 format = '(' + str(cell.value) + ')'
            #                 data_1.update({
            #                     format: ('E' + str(count))
            #                 })
            # date_from = datetime.strptime(options['date']['date_from'], "%Y-%m-%d")
            # date_to = datetime.strptime(options['date']['date_to'], "%Y-%m-%d")
            # period_type = options['date']['period_type']
            # if period_type == 'custom':
            #     date_dif = date_to - date_from
            #     new_date_to = date_from - timedelta(days=1)
            #     new_date_from = new_date_to - date_dif
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'quarter':
            #     new_date_from = date_from + relativedelta(months=-3)
            #     new_date_to = date_to + relativedelta(months=-3)
            #     new_date_from = date_utils.start_of(new_date_from, 'month')
            #     new_date_to = date_utils.end_of(new_date_to, 'month')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'month':
            #     new_date_from = date_from + relativedelta(months=-1)
            #     new_date_to = date_to + relativedelta(months=-1)
            #     new_date_from = date_utils.start_of(new_date_from, 'month')
            #     new_date_to = date_utils.end_of(new_date_to, 'month')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'fiscalyear':
            #     new_date_from = date_from + relativedelta(years=-1)
            #     new_date_to = date_to + relativedelta(years=-1)
            #     new_date_from = date_utils.start_of(new_date_from, 'year')
            #     new_date_to = date_utils.end_of(new_date_to, 'year')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            #
            # ctx = self._set_context(options)
            # ctx.update({'no_format': True, 'print_mode': True, 'prefetch_fields': False})
            # lines2 = self.with_context(ctx)._get_lines(options)
            # for line2 in lines2:
            #     if line2['columns'][0]['name'] != '':
            #         for key in data_1:
            #             if key in line2['name']:
            #                 sheet[data_1[key]] = line2['columns'][0]['name']

        elif self.id == self.env.ref('advanced_vn_report.account_financial_report_pnl_b02').id:
            template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                'Báo cáo kết quả hoạt động kinh doanh.xlsx')
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook['Sheet1']
            rows = sheet.rows
            count = 0
            data = {}
            data_1 = {}
            for row in rows:
                count += 1
                if 9 <= count <= 26:
                    for cell in row:
                        if cell.value is not None and cell.coordinate == ('G' + str(count)):
                            format = '(' + str(cell.value) + ')'
                            data.update({
                                format: ('I' + str(count))
                            })
                            data_1.update({
                                format: ('J' + str(count))
                            })
            ctx = self._set_context(options)
            ctx.update({'no_format': True, 'print_mode': True, 'prefetch_fields': False})
            lines = self.with_context(ctx)._get_lines(options)
            if len(lines) > 0:
                for line in lines:
                    if len(line['columns']) > 1:
                        if line['columns'][0]['name'] != '':
                            for key in data:
                                if key in line['name']:
                                    sheet[data[key]] = line['columns'][0]['name']
                        if line['columns'][1]['name'] != '':
                            for key in data:
                                if key in line['name']:
                                    sheet[data_1[key]] = line['columns'][1]['name']
                    else:
                        raise UserError(_('Không có dữ liệu cột so sánh theo kỳ.'))

            # dau ky
            # count = 0
            # data_1 = {}
            # rows2 = sheet.rows
            # for row2 in rows2:
            #     count += 1
            #     if 9 <= count <= 26:
            #         for cell in row2:
            #             if cell.value is not None and cell.coordinate == ('G' + str(count)):
            #                 format = '(' + str(cell.value) + ')'
            #                 data_1.update({
            #                     format: ('J' + str(count))
            #                 })
            # date_from = datetime.strptime(options['date']['date_from'], "%Y-%m-%d")
            # date_to = datetime.strptime(options['date']['date_to'], "%Y-%m-%d")
            # period_type = options['date']['period_type']
            # if period_type == 'custom':
            #     date_dif = date_to - date_from
            #     new_date_to = date_from - timedelta(days=1)
            #     new_date_from = new_date_to - date_dif
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'quarter':
            #     new_date_from = date_from + relativedelta(months=-3)
            #     new_date_to = date_to + relativedelta(months=-3)
            #     new_date_from = date_utils.start_of(new_date_from, 'month')
            #     new_date_to = date_utils.end_of(new_date_to, 'month')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'month':
            #     new_date_from = date_from + relativedelta(months=-1)
            #     new_date_to = date_to + relativedelta(months=-1)
            #     new_date_from = date_utils.start_of(new_date_from, 'month')
            #     new_date_to = date_utils.end_of(new_date_to, 'month')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            # elif period_type == 'fiscalyear':
            #     new_date_from = date_from + relativedelta(years=-1)
            #     new_date_to = date_to + relativedelta(years=-1)
            #     new_date_from = date_utils.start_of(new_date_from, 'year')
            #     new_date_to = date_utils.end_of(new_date_to, 'year')
            #     options['date']['date_from'] = new_date_from.strftime("%Y-%m-%d")
            #     options['date']['date_to'] = new_date_to.strftime("%Y-%m-%d")
            #
            # ctx = self._set_context(options)
            # ctx.update({'no_format': True, 'print_mode': True, 'prefetch_fields': False})
            # lines2 = self.with_context(ctx)._get_lines(options)
            # for line2 in lines2:
            #     if line2['columns'][0]['name'] != '':
            #         for key in data_1:
            #             if key in line2['name']:
            #                 sheet[data_1[key]] = line2['columns'][0]['name']
            sheet['G3'] = current_year
            sheet['D5'] = self.env.company.name
            # sheet['D4'] = self.env.company.vat
        elif report_manager.report_name == 'account.generic.tax.report':
            template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                'Báo cáo Thuế.xlsx')
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook['Sheet1']
            lines = self._get_lines(options, line_id=line_id)
            dict_sale = {}
            dict_purchase = {}
            try:
                for line in lines:
                    if '(10.0)' in line['name'] and self.env['account.tax'].sudo().search(
                            [('id', '=', line['id'])]).type_tax_use == 'sale':
                        dict_sale['(10.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                    elif '(5.0)' in line['name'] and self.env['account.tax'].sudo().search(
                            [('id', '=', line['id'])]).type_tax_use == 'sale':
                        dict_sale['(5.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                    elif '(0.0)' in line['name'] and self.env['account.tax'].sudo().search(
                            [('id', '=', line['id'])]).type_tax_use == 'sale':
                        dict_sale['(0.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                    elif '(10.0)' in line['name'] and self.env['account.tax'].sudo().search(
                            [('id', '=', line['id'])]).type_tax_use == 'purchase':
                        dict_purchase['(10.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                    elif '(5.0)' in line['name'] and self.env['account.tax'].sudo().search(
                            [('id', '=', line['id'])]).type_tax_use == 'purchase':
                        dict_purchase['(5.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                    elif '(0.0)' in line['name'] and self.env['account.tax'].sudo().search(
                            [('id', '=', line['id'])]).type_tax_use == 'purchase':
                        dict_purchase['(0.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                dau_phan_cach = self.env['res.lang'].sudo().search([('code', '=', self.env.user.lang)]).thousands_sep
                sheet['F31'] = float(
                    dict_sale["(10.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '')) if '(10.0)' in dict_sale and \
                                                                                            dict_sale["(10.0)"] else 0
                sheet['H31'] = float(
                    dict_sale["(10.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '')) if '(10.0)' in dict_sale and \
                                                                                            dict_sale["(10.0)"] else 0
                sheet['F30'] = float(
                    dict_sale["(5.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '')) if '(5.0)' in dict_sale and \
                                                                                           dict_sale["(5.0)"] else 0
                sheet['H30'] = float(
                    dict_sale["(5.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '')) if '(5.0)' in dict_sale and \
                                                                                           dict_sale["(5.0)"] else 0
                sheet['F29'] = float(
                    dict_sale["(0.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '')) if '(0.0)' in dict_sale and \
                                                                                           dict_sale["(0.0)"] else 0
                sheet['G29'] = float(
                    dict_sale["(0.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '')) if '(0.0)' in dict_sale and \
                                                                                           dict_sale["(0.0)"] else 0
                tong_von = 0
                tong_thue = 0
                for key in dict_purchase:
                    tong_von += float(dict_purchase[key][0].replace(' ₫', '').replace(dau_phan_cach, ''))
                    tong_thue += float(dict_purchase[key][1].replace(' ₫', '').replace(dau_phan_cach, ''))
                sheet['F24'] = tong_von
                sheet['H24'] = tong_thue
            except Exception as e:
                raise UserError("Không thể xuất báo cáo.")
        elif report_manager.report_name == 'account.assets.report':
            template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                'Báo cáo phân bổ.xlsx')
            workbook = openpyxl.load_workbook(template_path)
            sheet = workbook['Sheet1']
            ctx = self._set_context(options)
            date_from = datetime.strptime(ctx['date_from'], '%Y-%m-%d')
            date_to = datetime.strptime(ctx['date_to'], '%Y-%m-%d')
            if date_from.year != date_to.year:
                raise UserError('Chỉ chọn khoảng thời gian trong cùng năm')
            lines = self._get_lines(options, line_id=line_id)
            row = 9
            col = 1
            stt = 1
            # thoi gian bao cao
            sheet['G6'] = "Từ ngày " + date_from.strftime("%d/%m/%Y") + " đến ngày " + date_to.strftime("%d/%m/%Y")
            # ten cot gia ti luy ke theo context
            sheet['Y8'] = "Luỹ kế từ ngày " + date_from.strftime("%d/%m/%Y") + " đến ngày " + date_to.strftime(
                "%d/%m/%Y")
            tong_nguyen_gia = 0
            tong_phan_bo_thang = 0
            tong_luy_ke = 0
            tong_tong_luy_ke = 0
            tong_phan_bo_t1 = 0
            tong_phan_bo_t2 = 0
            tong_phan_bo_t3 = 0
            tong_phan_bo_t4 = 0
            tong_phan_bo_t5 = 0
            tong_phan_bo_t6 = 0
            tong_phan_bo_t7 = 0
            tong_phan_bo_t8 = 0
            tong_phan_bo_t9 = 0
            tong_phan_bo_t10 = 0
            tong_phan_bo_t11 = 0
            tong_phan_bo_t12 = 0
            tong_tong_gia_tri_con_lai = 0
            global_style = copy(sheet['A14']._style)
            if len(lines) > 45:
                so_dong_them = len(lines) - 45
                start_insert_row = 53
                while so_dong_them > 0:
                    sheet.insert_rows(start_insert_row)
                    for coltmp in range(28):
                        sheet.cell(row=start_insert_row, column=coltmp + 1)._style = global_style
                    start_insert_row = start_insert_row + 1
                    so_dong_them -= 1
            for line in lines:
                tong_gia_tri_con_lai = 0 if line['id'] == 'total' else line['columns'][-1]['no_format_name']
                if tong_gia_tri_con_lai:
                    tong_tong_gia_tri_con_lai += float(tong_gia_tri_con_lai)
                account_asset = False
                method = ''
                number_invoice = ''
                quantity = 1
                time_use = False
                nguyen_gia = 0
                ngay_mua = ''
                luy_ke = 0
                phan_bo_thang = 0
                phan_bo_t1 = 0
                phan_bo_t2 = 0
                phan_bo_t3 = 0
                phan_bo_t4 = 0
                phan_bo_t5 = 0
                phan_bo_t6 = 0
                phan_bo_t7 = 0
                phan_bo_t8 = 0
                phan_bo_t9 = 0
                phan_bo_t10 = 0
                phan_bo_t11 = 0
                phan_bo_t12 = 0
                if '_' in line['id']:
                    account_asset = self.env['account.asset'].search([('id', '=', line['id'][3:])])
                    if account_asset:
                        if account_asset.method == 'linear':
                            phan_bo_thang = account_asset.depreciation_move_ids[-1].amount_total
                            tong_phan_bo_thang += phan_bo_thang
                        if account_asset.depreciation_move_ids:
                            ngay_mua = account_asset.acquisition_date.strftime("%d/%m/%Y")
                            for asset_line in account_asset.depreciation_move_ids:
                                if asset_line.date.month == 1 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t1 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                for line in asset_line.line_ids:
                                                    account = line.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t1 += aline.debit - aline.credit
                                                    phan_bo_t1 = abs(phan_bo_t1)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t1
                                elif asset_line.date.month == 2 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t2 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t2 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t2 += aline.debit - aline.credit
                                                    phan_bo_t2 = abs(phan_bo_t2)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t2
                                elif asset_line.date.month == 3 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t3 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t3 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t3 += aline.debit - aline.credit
                                                    phan_bo_t3 = abs(phan_bo_t3)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t3
                                elif asset_line.date.month == 4 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t4 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t4 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t4 += aline.debit - aline.credit
                                                    phan_bo_t4 = abs(phan_bo_t4)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t4
                                elif asset_line.date.month == 5 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t5 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t5 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t5 += aline.debit - aline.credit
                                                    phan_bo_t5 = abs(phan_bo_t5)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t5
                                elif asset_line.date.month == 6 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t6 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t6 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t6 += aline.debit - aline.credit
                                                    phan_bo_t6 = abs(phan_bo_t6)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t6
                                elif asset_line.date.month == 7 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t7 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t7 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t7 += aline.debit - aline.credit
                                                    phan_bo_t7 = abs(phan_bo_t7)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t7
                                elif asset_line.date.month == 8 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t8 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t8 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t8 += aline.debit - aline.credit
                                                    phan_bo_t8 = abs(phan_bo_t8)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t8
                                elif asset_line.date.month == 9 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t9 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t9 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t9 += aline.debit - aline.credit
                                                    phan_bo_t9 = abs(phan_bo_t9)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t9
                                elif asset_line.date.month == 10 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t10 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t10 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t10 += aline.debit - aline.credit
                                                    phan_bo_t10 = abs(phan_bo_t10)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t10
                                elif asset_line.date.month == 11 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t11 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t11 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t11 += aline.debit - aline.credit
                                                    phan_bo_t11 = abs(phan_bo_t11)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t11
                                elif asset_line.date.month == 12 and date_from.year <= asset_line.date.year <= date_to.year:
                                    phan_bo_t12 = asset_line.amount_total
                                    if asset_line.asset_remaining_value == 0:
                                        if asset_line.asset_depreciated_value <= asset_line.amount_total:
                                            if asset_line.line_ids:
                                                phan_bo_t12 = 0
                                                for asline in asset_line.line_ids:
                                                    account = asline.account_id
                                                    for aline in asset_line.line_ids:
                                                        if aline.account_id == account:
                                                            phan_bo_t12 += aline.debit - aline.credit
                                                    phan_bo_t12 = abs(phan_bo_t12)
                                                    break
                                    if date_from.month <= asset_line.date.month <= date_to.month:
                                        luy_ke += phan_bo_t12
                        nguyen_gia = account_asset.original_value - account_asset.salvage_value
                        if account_asset.method_period == '1':
                            time_use = account_asset.method_number
                        if account_asset.method_period == '12':
                            time_use = account_asset.method_number * 12
                        if account_asset.original_move_line_ids:
                            quantity = 0
                            n = 0
                            for rec in account_asset.original_move_line_ids:
                                quantity += rec.quantity
                                if rec.move_id:
                                    n = n + 1
                            for rec in account_asset.original_move_line_ids:
                                n = n - 1
                                if n > 0:
                                    number_invoice += rec.move_id.name + ', '
                                else:
                                    number_invoice += rec.move_id.name
                col = 1
                sheet.cell(row=row, column=col).value = stt
                col += 1
                # ma tai san
                sheet.cell(row=row,
                           column=col).value = account_asset.asset_code if account_asset and account_asset.asset_code else None
                col += 1
                # ten tai san
                sheet.cell(row=row, column=col).value = account_asset.name if account_asset and line['level'] > 0 else \
                    line['name'] if "name" in line and line['level'] == 0 else None
                col += 1
                # so hoa don
                sheet.cell(row=row, column=col).value = number_invoice if line['level'] > 0 else None
                col += 1
                # so luong
                sheet.cell(row=row, column=col).value = quantity if line['level'] > 0 else None
                col += 1
                # nguyen gia TS
                tong_nguyen_gia += nguyen_gia
                sheet.cell(row=row, column=col).value = nguyen_gia if line['level'] > 0 else tong_nguyen_gia if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                # ngay mua
                sheet.cell(row=row, column=col).value = ngay_mua if line['level'] > 0 else None
                col += 1
                # ngay dua vao su dung
                sheet.cell(row=row, column=col).value = account_asset.first_depreciation_date.strftime(
                    "%d/%m/%Y") if account_asset and account_asset.first_depreciation_date else None
                col += 1
                # thoi gian su dung
                sheet.cell(row=row, column=col).value = time_use if line['level'] > 0 else None
                col += 1
                # muc phan bo thang

                # sheet.cell(row=row, column=col).value = phan_bo_thang if line['level'] > 0 else tong_phan_bo_thang if \
                #     line['id'] == 'total' else None
                if line['level'] > 0:
                    if phan_bo_thang > 0:
                        sheet.cell(row=row, column=col).value = phan_bo_thang
                if line['id'] == 'total':
                    sheet.cell(row=row, column=col).value = tong_phan_bo_thang
                col += 1
                # TK no
                sheet.cell(row=row,
                           column=col).value = account_asset.account_depreciation_expense_id.code if account_asset and \
                                                                                                     line[
                                                                                                         'level'] > 0 else None
                col += 1
                # TK co
                sheet.cell(row=row, column=col).value = account_asset.account_depreciation_id.code if account_asset and \
                                                                                                      line[
                                                                                                          'level'] > 0 else None
                col += 1
                tong_phan_bo_t1 += phan_bo_t1
                sheet.cell(row=row, column=col).value = phan_bo_t1 if line['level'] > 0 else tong_phan_bo_t1 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t2 += phan_bo_t2
                sheet.cell(row=row, column=col).value = phan_bo_t2 if line['level'] > 0 else tong_phan_bo_t2 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t3 += phan_bo_t3
                sheet.cell(row=row, column=col).value = phan_bo_t3 if line['level'] > 0 else tong_phan_bo_t3 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t4 += phan_bo_t4
                sheet.cell(row=row, column=col).value = phan_bo_t4 if line['level'] > 0 else tong_phan_bo_t4 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t5 += phan_bo_t5
                sheet.cell(row=row, column=col).value = phan_bo_t5 if line['level'] > 0 else tong_phan_bo_t5 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t6 += phan_bo_t6
                sheet.cell(row=row, column=col).value = phan_bo_t6 if line['level'] > 0 else tong_phan_bo_t6 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t7 += phan_bo_t7
                sheet.cell(row=row, column=col).value = phan_bo_t7 if line['level'] > 0 else tong_phan_bo_t7 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t8 += phan_bo_t8
                sheet.cell(row=row, column=col).value = phan_bo_t8 if line['level'] > 0 else tong_phan_bo_t8 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t9 += phan_bo_t9
                sheet.cell(row=row, column=col).value = phan_bo_t9 if line['level'] > 0 else tong_phan_bo_t9 if line[
                                                                                                                    'id'] == 'total' else None
                col += 1
                tong_phan_bo_t10 += phan_bo_t10
                sheet.cell(row=row, column=col).value = phan_bo_t10 if line['level'] > 0 else tong_phan_bo_t10 if line[
                                                                                                                      'id'] == 'total' else None
                col += 1
                tong_phan_bo_t11 += phan_bo_t11
                sheet.cell(row=row, column=col).value = phan_bo_t11 if line['level'] > 0 else tong_phan_bo_t11 if line[
                                                                                                                      'id'] == 'total' else None
                col += 1
                tong_phan_bo_t12 += phan_bo_t12
                sheet.cell(row=row, column=col).value = phan_bo_t12 if line['level'] > 0 else tong_phan_bo_t12 if line[
                                                                                                                      'id'] == 'total' else None
                col += 1
                tong_luy_ke += luy_ke
                luy_ke_nam = phan_bo_t1 + phan_bo_t2 + phan_bo_t3 + phan_bo_t4 + phan_bo_t5 + phan_bo_t6 + phan_bo_t7 + phan_bo_t8 + phan_bo_t9 + phan_bo_t10 + phan_bo_t11 + phan_bo_t12
                tong_tong_luy_ke += luy_ke_nam
                sheet.cell(row=row, column=col).value = luy_ke if line['level'] > 0 else tong_luy_ke if line[
                                                                                                            'id'] == 'total' else None
                col += 1
                sheet.cell(row=row, column=col).value = luy_ke_nam if line['level'] > 0 else tong_tong_luy_ke if line[
                                                                                                                     'id'] == 'total' else None
                col += 1
                sheet.cell(row=row, column=col).value = tong_gia_tri_con_lai if line[
                                                                                    'level'] > 0 else tong_tong_gia_tri_con_lai if \
                    line['id'] == 'total' else None
                row += 1
                stt += 1

        elif report_manager.report_name == 'account.general.ledger':
            template_path1 = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                                 'Mẫu-sổ-nhật-ký-chung.xlsx')
            workbook = openpyxl.load_workbook(template_path1)
            sheet = workbook['Sheet1']
            global_style = copy(sheet['A11']._style)
            global_style_1 = copy(sheet['E11']._style)
            global_style_2 = copy(sheet['C11']._style)
            lines = self._get_lines(options, line_id=line_id)
            date_to = datetime.strptime(options['date'].get('date_from'), '%Y-%m-%d').strftime("%d/%m/%Y")
            date_from = datetime.strptime(options['date'].get('date_to'), '%Y-%m-%d').strftime("%d/%m/%Y")
            sheet['A5'] = 'Từ ngày ' + date_to + ' đến ngày ' + date_from

            # lay ra 1 dict chua key: account, value account move line
            data = {}
            list_children = []
            parent_line_id = None
            for line in lines:
                if 'level' in line and line['level'] == 2 and 'unfolded' in line and line['unfolded']:
                    if not parent_line_id:
                        parent_line_id = lines[0]['id']
                    else:
                        parent_line_id = line['id']
                    data[parent_line_id] = []

                else:
                    if 'level' in line and line['level'] == 4 and line['parent_id'] == parent_line_id:
                        list_children = data[parent_line_id]
                        list_children.append(line)
                        data[parent_line_id] = list_children

            for y in range(0, len(lines)):
                level = lines[y].get('level')
                unfolded = lines[y].get('unfolded')

                # write the first column, with a specific style to manage the indentation
                if level == 2 and unfolded:
                    new_sheet = workbook.copy_worksheet(sheet)
                    new_sheet.title = lines[y]['title_hover']
                    title_hover = lines[y].get('title_hover').split(' ')[0]
                    list_caret = 0
                    if lines[y]['id'] in data:
                        list_caret = len(data[lines[y]['id']])
                    if list_caret >= 10:
                        so_dong_them = list_caret - 10
                        start_insert_row = 20
                        while so_dong_them > 0:
                            new_sheet.insert_rows(start_insert_row)
                            for coltmp in range(6):
                                if coltmp in range(4) and coltmp != 2:
                                    new_sheet.cell(row=start_insert_row, column=coltmp + 1)._style = global_style
                                elif coltmp in range(4, 6):
                                    new_sheet.cell(row=start_insert_row, column=coltmp + 1)._style = global_style_1
                                elif coltmp == 2:
                                    new_sheet.cell(row=start_insert_row, column=coltmp + 1)._style = global_style_2
                            start_insert_row = start_insert_row + 1
                            so_dong_them -= 1
                    max_length = 0
                    y_offset = 11
                total_end = 0
                total_debit_end = 0
                total_credit_end = 0
                dau_phan_cach = self.env['res.lang'].sudo().search([('code', '=', self.env.user.lang)]).thousands_sep
                if lines[y].get('caret_options'):
                    # lay gtri tung dong cua account move
                    cell_type, cell_value = self._get_cell_type_value(lines[y])
                    # neu cell type = date thi fill vao cot C
                    if cell_type == 'date':
                        new_sheet['B' + str(y_offset)] = cell_value.strftime(
                            "%d/%m/%Y")
                    else:
                        new_sheet['A' + str(y_offset)] = cell_value
                        # lấy ra id account_move
                        if str(lines[y]['id']).isnumeric():
                            # Tìm account_move_line
                            acc_move_line = self.env['account.move.line'].sudo().search(
                                [('id', '=', int(lines[y]['id']))])
                            # Lấy ra các account_move
                            acc_move = self.env['account.move'].sudo().search([('id', '=', acc_move_line.move_id.id)])
                            list_account = []
                            for account in acc_move.vn_line_ids:
                                # Lấy all tk trong account_move
                                if acc_move.vn_line_ids:
                                    if account.account_id.code:
                                        if account.account_id.code == title_hover:
                                            list_account.append(account.related_account_id.code)
                                            list_account = list(set(list_account))
                                            # Fill tài khoản vào Excel
                                            new_sheet['D' + str(y_offset)] = ', '.join(list_account)
                                            # print(','.join(list_account))
                    for x in range(1, len(lines[y]['columns']) + 1):
                        cell_type, cell_value = self._get_cell_type_value(lines[y]['columns'][x - 1])
                        if cell_type == 'date':
                            new_sheet.cell(row=y_offset, column=lines[y].get('colspan', 2),
                                           value=cell_value.strftime(
                                               "%d/%m/%Y"))
                        if x == 2:
                            if len(str(cell_value)) > max_length:
                                max_length = len(cell_value)
                            pass
                            new_sheet.column_dimensions[get_column_letter(x + 1)].width = max_length
                            new_sheet.cell(row=y_offset, column=lines[y].get('colspan', 3),
                                           value=cell_value)
                        # if x == 5:
                        #     new_sheet.cell(row=y_offset, column=lines[y].get('colspan', 4),
                        #                    value=cell_value)

                        if x == 6:
                            number = cell_value.replace('.00 ₫', '')
                            number = number.replace(' ₫', '')
                            number = number.replace('₫', '')
                            number = number.replace(',', '')
                            result = number.replace('.', '')
                            if result == '':
                                continue
                            cell_value = float(result)
                            new_sheet.cell(row=y_offset, column=lines[y].get('colspan', 5),
                                           value=cell_value)

                        if x == 7:
                            number = cell_value.replace('.00 ₫', '')
                            number = number.replace(' ₫', '')
                            number = number.replace('₫', '')
                            number = number.replace(',', '')
                            result = number.replace('.', '')
                            if result == '':
                                continue
                            cell_value = float(result)
                            new_sheet.cell(row=y_offset, column=lines[y].get('colspan', 6),
                                           value=cell_value)
                        else:
                            pass
                    y_offset += 1

                # lay so du dau ky
                elif lines[y].get('class') == 'o_account_reports_initial_balance':
                    if lines[y]['columns']:
                        new_sheet.cell(row=10, column=4,
                                       value=float(lines[y]['columns'][1]['name'].replace(' ₫', '').replace(',', '')))
                        new_sheet.cell(row=10, column=5,
                                       value=float(lines[y]['columns'][2]['name'].replace(' ₫', '').replace(',', '')))
                        new_sheet.cell(row=10, column=6,
                                       value=float(lines[y]['columns'][3]['name'].replace(' ₫', '').replace(',', '')))
                # lay total
                elif lines[y].get('class') == 'o_account_reports_domain_total':
                    # cell_type, cell_value = self._get_cell_type_value(lines[y]['columns'])
                    if lines[y]['columns']:
                        if list_caret >= 10:
                            new_sheet['E' + str(list_caret + 11)] = '=SUM(E11:E' + str(list_caret + 10) + ')'
                            new_sheet['F' + str(list_caret + 11)] = '=SUM(F11:F' + str(list_caret + 10) + ')'
                            new_sheet['D' + str(list_caret + 11)] = '=SUM(E11:E' + str(
                                list_caret + 10) + ')' + '-' + 'SUM(F11:F' + str(list_caret + 10) + ') ' + '+' + 'D10 '

                            # new_sheet.cell(row=list_caret + 11, column=4,
                            #                value=lines[y]['columns'][1]['name'])
                            # new_sheet.cell(row=list_caret + 11, column=5,
                            #                value=lines[y]['columns'][2]['name'])
                            # new_sheet.cell(row=list_caret + 11, column=6,
                            #                value=lines[y]['columns'][3]['name'])

                        else:
                            new_sheet['E21'] = '=SUM(E11:E20)'
                            new_sheet['F21'] = '=SUM(F11:F20)'
                            new_sheet['D21'] = '=SUM(E11:E20) - SUM(F11:F20) + D10'
                            # new_sheet.cell(row=21, column=4,
                            #                value=lines[y]['columns'][1]['name'])
                            # new_sheet.cell(row=21, column=5,
                            #                value=lines[y]['columns'][2]['name'])
                            # new_sheet.cell(row=21, column=6,
                            #                value=lines[y]['columns'][3]['name'])
            workbook.remove(sheet)
            if len(workbook.worksheets) > 0:
                pass
            else:
                raise UserError('Báo cáo không có dữ liệu.')
            #     workbook.remove(sheet)
            #     if workbook['worksheets']:
            #         print('test')

        else:
            raise UserError('Không hỗ trợ loại báo cáo này')
        workbook.save(output)
        # out = base64.encodestring(content.getvalue())
        # workbook = xlsxwriter.Workbook(output, {'in_memory': True})
        # sheet = workbook.add_worksheet(self._get_report_name()[:31])

        workbook.close()
        output.seek(0)
        generated_file = output.read()
        output.close()

        return generated_file

    def action_view_account_move_tree(self, options, params=None):
        active_id = int(params.get('id').split('_')[-1])
        line = self.env['account.asset'].browse(active_id)
        action = self.env.ref('account.action_move_journal_line').read()[0]
        action['domain'] = [('id', 'in', line.depreciation_move_ids.ids)]
        action['view_mode'] = 'list,form'
        action['views'] = [(self.env.ref('account.view_move_tree').id, 'list'), (False, 'form')]
        action['context'] = {'create': False, 'edit': True}
        return action

    def open_journal_items(self, options, params):
        result = super(AccountReport, self).open_journal_items(options, params)
        # search account_move_line with new viet nam domain first
        is_our_vn_report = False
        external_data = self.env['ir.model.data'].search(
            [('model', '=', 'account.financial.html.report'), ('res_id', '=', self.id)])
        if external_data:
            if external_data.name == 'account_financial_report_b01':
                is_our_vn_report = True
                if 'domain' in result and len(result['domain']) == 3:
                    result['domain'] = [result['domain'][2]]
            elif external_data.name == 'account_financial_report_pnl_b02':
                is_our_vn_report = True
            elif external_data.name == 'account_financial_report_cash_flow_statement_b03':
                is_our_vn_report = True
        if is_our_vn_report:
            if 'domain' in params:
                new_domain = self.convert_domain_not_compare(safe_eval(params['domain']))
                vn_account_move_line = self.env['vn.account.move.line'].search(new_domain)
                account_move_ids = [e.account_move_line_backup.id for e in vn_account_move_line]
                result['domain'] = [('id', 'in', account_move_ids)] + result['domain']
        return result


    def convert_domain_not_compare(self, domain):
        for n, i in enumerate(domain):
            if type(i) == tuple:
                list_tuple = []
                for j in i:
                    if type(j) == str:
                        if "_compare" in j:
                            j = j.replace("_compare", "")
                            list_tuple.append(j)
                        else:
                            list_tuple.append(j)
                    else:
                        list_tuple.append(j)
                domain[n] = tuple(list_tuple)
        return domain
