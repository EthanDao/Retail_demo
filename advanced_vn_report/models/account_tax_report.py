import base64
import io
import logging
import os
from datetime import date

import openpyxl

from odoo import fields, models
from odoo.exceptions import UserError
from odoo.modules import get_module_resource

_logger = logging.getLogger(__name__)

# from BeautifulSoup import BeautifulStoneSoup as Soup
try:
    from openpyxl import load_workbook
    from openpyxl import Workbook
    from openpyxl.utils.exceptions import IllegalCharacterError
except ImportError:
    _logger.debug(
        'Cannot import "openpyxl". Please make sure it is installed.')


class AccountTaxReport(models.Model):
    _inherit = "account.generic.tax.report"
    _name = "account.xml.tax.report"
    _rec_name = 'file_name'

    file = fields.Binary(string="File kê khai thuế (XML)", copy=False)
    file_name = fields.Char(string="Tờ khai thuế")
    kyKKhai = fields.Char(string="Kỳ khai")
    kyKKhaiTuNgay = fields.Date(string="Kỳ khai từ ngày")
    kyKKhaiDenNgay = fields.Date(string="Kỳ khai đến ngày")
    maCQTNoiNop = fields.Char(string="Mã cơ quan thuế nơi nộp")
    tenCQTNoiNop = fields.Char(string="Tên cơ quan thuế nơi nộp")
    nguoiKy = fields.Char(string='Người khai')
    ngayKy = fields.Date(string="Ngày ký")
    nganhNgheKD = fields.Char(string="Ngành nghề kinh doanh")
    company_currency_id = fields.Many2one('res.currency', readonly=True, default=lambda x: x.env.company.currency_id)

    def _get_default_ct22(self):
        report = self.env['account.xml.tax.report'].search([], order='id desc', limit=1)
        if report:
            return report.ct_43
        else:
            return 0

    ct_22 = fields.Monetary(string="Thuế GTGT còn được khấu trừ kỳ trước chuyển sang [22]", currency_field="company_currency_id", default=_get_default_ct22)
    ct_25 = fields.Monetary(string="Tổng số thuế GTGT được khấu trừ kỳ này [25]", currency_field="company_currency_id")
    ct_37 = fields.Monetary(string="Điều chỉnh giảm thuế GTGT còn được khấu trừ của các kỳ trước [37]", currency_field="company_currency_id")
    ct_38 = fields.Monetary(string="Điều chỉnh tăng thuế GTGT còn được khấu trừ của các kỳ trước [38]", currency_field="company_currency_id")
    ct_39 = fields.Monetary(string="Thuế GTGT đã nộp ở địa phương khác của hoạt động kinh doanh xây dựng, lắp đặt, bán hàng, bất động sản ngoại tỉnh [39]", currency_field="company_currency_id")
    ct_40b = fields.Monetary(string="Thuế GTGT mua vào của dự án đầu tư được bù trừ với thuế GTGT còn phải nộp của hoạt động sản xuất kinh doanh cùng kỳ tính thuế [40b]",
                             currency_field="company_currency_id")
    ct_42 = fields.Monetary(string="Tổng số thuế GTGT đề nghị hoàn [42]", currency_field="company_currency_id")
    ct_43 = fields.Monetary(string="Thuế GTGT còn được khấu trừ chuyển kỳ sau ([43]=[41]-[42])", currency_field="company_currency_id")

    def export_xml_data(self):
        # template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
        #                                     'TỜ KHAI THUẾ GIÁ TRỊ GIA TĂNG.xml')
        # file_content = open(template_path, 'w')
        file_content = open("TỜ KHAI THUẾ GIÁ TRỊ GIA TĂNG.xml", "w")
        content = """<?xml version="1.0" encoding="UTF-8"?>
<HSoThueDTu xmlns="http://kekhaithue.gdt.gov.vn/TKhaiThue" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance">
    <HSoKhaiThue id="ID_18">
        <TTinChung>
            <TTinDVu>
                <maDVu>HTKK</maDVu>
                <tenDVu>HTKK</tenDVu>
                <pbanDVu>3.8.6</pbanDVu>
                <ttinNhaCCapDVu/>
            </TTinDVu>
            <TTinTKhaiThue>
                <TKhaiThue>
                    <maTKhai>01</maTKhai>
                    <tenTKhai>TỜ KHAI THUẾ GIÁ TRỊ GIA TĂNG (Mẫu số 01/GTGT)</tenTKhai>
                    <moTaBMau>(Ban hành kèm theo Thông tư số 26/2015/TT-BTC ngày 27/02/2015 của Bộ Tài chính)</moTaBMau>
                    <pbanTKhaiXML>2.1.2</pbanTKhaiXML>
                    <loaiTKhai>C</loaiTKhai>
                    <soLan>0</soLan>
                    <KyKKhaiThue>
                        <kieuKy>M</kieuKy>
                        <kyKKhai>var_kyKKhai</kyKKhai>
                        <kyKKhaiTuNgay>var_kyKKhaiTuNgay</kyKKhaiTuNgay>
                        <kyKKhaiDenNgay>var_kyKKhaiDenNgay</kyKKhaiDenNgay>
                        <kyKKhaiTuThang/>
                        <kyKKhaiDenThang/>
                    </KyKKhaiThue>
                    <maCQTNoiNop>var_maCQTNoiNop</maCQTNoiNop>
                    <tenCQTNoiNop>var_tenCQTNoiNop</tenCQTNoiNop>
                    <ngayLapTKhai>var_ngayLapTKhai</ngayLapTKhai>
                    <GiaHan>
                        <maLyDoGiaHan/>
                        <lyDoGiaHan/>
                    </GiaHan>
                    <nguoiKy>var_nguoiKy</nguoiKy>
                    <ngayKy>var_ngayKy</ngayKy>
                    <nganhNgheKD>var_nganhNgheKD</nganhNgheKD>
                </TKhaiThue>
                <NNT>
                    <mst>var_mst</mst>
                    <tenNNT>var_tenNNT</tenNNT>
                    <dchiNNT>var_dchiNNT</dchiNNT>
                    <phuongXa/>
                    <maHuyenNNT/>
                    <tenHuyenNNT>var_tenHuyenNNT</tenHuyenNNT>
                    <maTinhNNT/>
                    <tenTinhNNT>var_tenTinhNNT</tenTinhNNT>
                    <dthoaiNNT>var_dthoaiNNT</dthoaiNNT>
                    <faxNNT>var_faxNNT</faxNNT>
                    <emailNNT>var_emailNNT</emailNNT>
                </NNT>
            </TTinTKhaiThue>
        </TTinChung>
        <CTieuTKhaiChinh>
            <tieuMucHachToan>1701</tieuMucHachToan>
            <ct21>var_ct21</ct21>
            <ct22>var_ct22</ct22>
            <GiaTriVaThueGTGTHHDVMuaVao>
                <ct23>var_ct23</ct23>
                <ct24>var_ct24</ct24>
            </GiaTriVaThueGTGTHHDVMuaVao>
            <ct25>var_ct25</ct25>
            <ct26>var_ct26</ct26>
            <HHDVBRaChiuThueGTGT>
                <ct27>var_ct27</ct27>
                <ct28>var_ct28</ct28>
            </HHDVBRaChiuThueGTGT>
            <ct29>var_ct29</ct29>
            <HHDVBRaChiuTSuat5>
                <ct30>var_ct30</ct30>
                <ct31>var_ct31</ct31>
            </HHDVBRaChiuTSuat5>
            <HHDVBRaChiuTSuat10>
                <ct32>var_ct32</ct32>
                <ct33>var_ct33</ct33>
            </HHDVBRaChiuTSuat10>
            <HHDVBRaKhongTinhThue>
                <ct32a>var_ct32a</ct32a>
            </HHDVBRaKhongTinhThue>
            <TongDThuVaThueGTGTHHDVBRa>
                <ct34>var_ct34</ct34>
                <ct35>var_ct35</ct35>
            </TongDThuVaThueGTGTHHDVBRa>
            <ct36>var_ct36</ct36>
            <ct37>var_ct37</ct37>
            <ct38>var_ct38</ct38>
            <ct39>var_ct39</ct39>
            <ct40a>var_ct40a</ct40a>
            <ct40b>var_ct40b</ct40b>
            <ct40>var_ct40</ct40>
            <ct41>var_ct41</ct41>
            <ct42>var_ct42</ct42>
            <ct43>var_ct43</ct43>
        </CTieuTKhaiChinh>
        <PLuc>
            <PL01_6_GTGT>
                <ct06>0</ct06>
                <ct07>0</ct07>
                <PBoGTGTKhongHTKT>
                    <PBoThueGTGTChoDiaPhuong>
                        <PhanBoGTGT id="ID_1">
                            <ct09>0</ct09>
                            <ct10>0</ct10>
                            <ct11_maCT>0</ct11_maCT>
                            <ct11_ct>var_ct11_ct</ct11_ct>
                            <ct11_maCCT>0</ct11_maCCT>
                            <ct11_cct>var_ct11_cct</ct11_cct>
                            <DThuSPSXCuaCSSXPhuThuoc>
                                <ct12>0</ct12>
                                <ct13>0</ct13>
                                <ct14>0</ct14>
                            </DThuSPSXCuaCSSXPhuThuoc>
                            <ct15>0</ct15>
                            <ct17>0</ct17>
                        </PhanBoGTGT>
                    </PBoThueGTGTChoDiaPhuong>
                    <ct16>0</ct16>
                    <ct18>0</ct18>
                    <ct19>0</ct19>
                    <ct20>0</ct20>
                </PBoGTGTKhongHTKT>
            </PL01_6_GTGT>
        </PLuc>
    </HSoKhaiThue>
</HSoThueDTu>"""

        options = {'unfolded_lines': [],
                   'date': {'string': '2020', 'period_type': 'month', 'mode': 'range', 'date_from': str(self.kyKKhaiTuNgay.strftime("%Y-%m-%d")),
                            'date_to': str(self.kyKKhaiDenNgay.strftime("%Y-%m-%d")), 'filter': 'this_month', 'strict_range': True},
                   'comparison': {'filter': 'no_comparison', 'number_period': 1, 'date_from': '', 'date_to': '', 'periods': []}, 'all_entries': False, 'unposted_in_period': True}
        # options = {
        #     'unfolded_lines': [],
        #     'date': {'string': 'May 2020', 'period_type': 'month', 'mode': 'range', 'date_from': '2020-05-01', 'date_to': '2020-05-31', 'filter': 'this_month', 'strict_range': True},
        #     'comparison': {'filter': 'no_comparison', 'number_period': 1, 'date_from': '', 'date_to': '', 'periods': []},
        #     'all_entries': False
        # }
        filter_date = {'mode': 'range', 'filter': 'last_month'}
        filter_comparison = {'date_from': '', 'date_to': '', 'filter': 'no_comparison', 'number_period': 1}
        account_report = self.env['account.generic.tax.report']
        account_report = account_report.with_context(account_report._set_context(options))
        account_report = account_report.with_context(model='account.generic.tax.report')
        account_report.filter_date = filter_date
        account_report.filter_comparison = filter_comparison
        lines = account_report._get_lines(options)
        ct_21 = ''  # Không phát sinh hoạt động mua, bán trong kỳ (đánh dấu "X")
        ct_22 = 0  # Thuế GTGT còn được khấu trừ kỳ trước chuyển sang
        ct_23 = 0  # Giá trị của hàng hoá, dịch vụ mua vào (chưa VAT) PO VAT
        ct_24 = 0  # thuế GTGT của hàng hoá, dịch vụ mua vào (VAT)
        ct_25 = 0  # Tổng số thuế GTGT được khấu trừ kỳ này
        ct_26 = 0  # Hàng hóa, dịch vụ bán ra không chịu thuế GTGT
        ct_27 = 0  # Hàng hóa, dịch vụ bán ra chịu thuế GTGT ([27]=[29]+[30]+[32]+[32a])
        ct_28 = 0  # Hàng hóa, dịch vụ bán ra chịu thuế GTGT ([28]=[31]+[33])
        ct_29 = 0  # Hàng hoá, dịch vụ bán ra chịu thuế suất 0%
        ct_30 = 0  # Hàng hoá, dịch vụ bán ra chịu thuế suất 5% (Chưa VAT)
        ct_31 = 0  # chịu thuế suất 5% (VAT)
        ct_32 = 0  # Hàng hoá, dịch vụ bán ra chịu thuế suất 10%(Chưa VAT)
        ct_32a = 0  # Hàng hoá, dịch vụ bán ra không tính thuế
        ct_33 = 0  # chịu thuế suất 10%(VAT)
        ct_34 = 0  # Tổng doanh thu của HHDV bán ra ([34]=[26]+[27]; [35]=[28])(Chưa VAT)
        ct_35 = 0  # Tổng thuế GTGT của HHDV bán ra ([34]=[26]+[27]; [35]=[28]) (VAT)
        ct_36 = 0  # Thuế GTGT phát sinh trong kỳ ([36]=[35]-[25])
        ct_37 = 0  # Điều chỉnh giảm
        ct_38 = 0  # Điều chỉnh tăng
        ct_39 = 0  # Thuế GTGT đã nộp ở địa phương khác của hoạt động kinh doanh xây dựng, lắp đặt, bán hàng, bất động sản ngoại tỉnh
        ct_40a = 0  # Thuế GTGT phải nộp của hoạt động sản xuất kinh doanh trong kỳ ([40a]=[36]-[22]+[37]-[38] - [39]≥ 0)
        ct_40b = 0  # Thuế GTGT mua vào của dự án đầu tư được bù trừ với thuế GTGT còn phải nộp của hoạt động sản xuất kinh doanh cùng kỳ tính thuế
        ct_40 = 0  # Thuế GTGT còn phải nộp trong kỳ ([40]=[40a]-[40b])
        ct_41 = 0  # Thuế GTGT chưa khấu trừ hết kỳ này (nếu [41]=[36]-[22]+[37]-[38]-[39]< 0)
        ct_42 = 0  # Tổng số thuế GTGT đề nghị hoàn
        ct_43 = 0  # Thuế GTGT còn được khấu trừ chuyển kỳ sau ([43]=[41]-[42])
        if lines:
            dict_sale = {}
            dict_purchase = {}
            for line in lines:
                if '(10.0)' in line['name'] and self.env['account.tax'].sudo().search([('id', '=', line['id'])]).type_tax_use == 'sale':
                    dict_sale['(10.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                elif '(5.0)' in line['name'] and self.env['account.tax'].sudo().search([('id', '=', line['id'])]).type_tax_use == 'sale':
                    dict_sale['(5.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                elif '(0.0)' in line['name'] and self.env['account.tax'].sudo().search([('id', '=', line['id'])]).type_tax_use == 'sale':
                    dict_sale['(0.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                elif '(10.0)' in line['name'] and self.env['account.tax'].sudo().search([('id', '=', line['id'])]).type_tax_use == 'purchase':
                    dict_purchase['(10.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                elif '(5.0)' in line['name'] and self.env['account.tax'].sudo().search([('id', '=', line['id'])]).type_tax_use == 'purchase':
                    dict_purchase['(5.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
                elif '(0.0)' in line['name'] and self.env['account.tax'].sudo().search([('id', '=', line['id'])]).type_tax_use == 'purchase':
                    dict_purchase['(0.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
            dau_phan_cach = self.env['res.lang'].sudo().search([('code', '=', self.env.user.lang)]).thousands_sep
            ct_32 = float(dict_sale["(10.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '') if '(10.0)' in dict_sale and dict_sale["(10.0)"] else 0)
            ct_33 = float(dict_sale["(10.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '') if '(10.0)' in dict_sale and dict_sale["(10.0)"] else 0)
            ct_30 = float(dict_sale["(5.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '') if '(5.0)' in dict_sale and dict_sale["(5.0)"] else 0)
            ct_31 = float(dict_sale["(5.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '') if '(5.0)' in dict_sale and dict_sale["(5.0)"] else 0)
            ct_29 = float(dict_sale["(0.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '') if '(0.0)' in dict_sale and dict_sale["(0.0)"] else 0)
            tong_von = 0
            in_invoices = self.env['account.move'].sudo().search(
                [('state', '=', 'posted'), ('type', '=', 'in_invoice'),
                 ('date', '<=', self.kyKKhaiDenNgay),
                 ('date', '>=', self.kyKKhaiTuNgay)])
            in_refunds = self.env['account.move'].sudo().search(
                [('state', '=', 'posted'), ('type', '=', 'in_refund'),
                 ('date', '<=', self.kyKKhaiDenNgay),
                 ('date', '>=', self.kyKKhaiTuNgay)])
            for in_invoice in in_invoices:
                tong_von += in_invoice.amount_untaxed
            for in_refund in in_refunds:
                tong_von -= in_refund.amount_untaxed
            tong_thue = 0
            for key in dict_purchase:
                # tong_von += int(dict_purchase[key][0].replace(' ₫', '').replace(dau_phan_cach, ''))
                tong_thue += float(dict_purchase[key][1].replace(' ₫', '').replace(dau_phan_cach, ''))
            ct_23 = tong_von
            ct_24 = tong_thue

            ct_22 = self.ct_22 if self.ct_22 else 0
            ct_25 = self.ct_25 if self.ct_25 else 0
            # start compute ct_26
            non_tax_total = 0
            invoices = self.env['account.move'].sudo().search(
                [('state', '=', 'posted'), ('type', '=', 'out_invoice'), ('date', '<=', self.kyKKhaiDenNgay),
                 ('date', '>=', self.kyKKhaiTuNgay)])
            out_refunds = self.env['account.move'].sudo().search(
                [('state', '=', 'posted'), ('type', '=', 'out_refund'), ('date', '<=', self.kyKKhaiDenNgay),
                 ('date', '>=', self.kyKKhaiTuNgay)])
            for invoice in invoices:
                if invoice.invoice_line_ids:
                    for inline in invoice.invoice_line_ids:
                        if not inline.tax_ids:
                            non_tax_total += inline.price_subtotal
            for out_refund in out_refunds:
                if out_refund.invoice_line_ids:
                    for inline in out_refund.invoice_line_ids:
                        if not inline.tax_ids:
                            non_tax_total -= inline.price_subtotal
            ct_26 = non_tax_total
            # end compute ct_26

            # start compute ct_25
            tax_25 = 0
            invoices_25 = self.env['account.move'].sudo().search(
                [('state', '=', 'posted'), ('type', '=', 'in_invoice'), ('nontaxable_income_source', '=', False),
                 ('date', '<=', self.kyKKhaiDenNgay),
                 ('date', '>=', self.kyKKhaiTuNgay)])
            invoices_25_refund = self.env['account.move'].sudo().search(
                [('state', '=', 'posted'), ('type', '=', 'in_refund'), ('nontaxable_income_source', '=', False),
                 ('date', '<=', self.kyKKhaiDenNgay),
                 ('date', '>=', self.kyKKhaiTuNgay)])
            for invoice_25 in invoices_25:
                tax_25 += invoice_25.amount_tax
            for invoice_25_re in invoices_25_refund:
                tax_25 -= invoice_25_re.amount_tax
            dt_0_10 = ct_29 + ct_32
            dt_khong_thue = ct_26
            ct_25 = round(tax_25 / (dt_0_10 + dt_khong_thue) * dt_0_10)
            # stop ompute ct_25

            ct_37 = self.ct_37 if self.ct_37 else 0
            ct_38 = self.ct_38 if self.ct_38 else 0
            ct_39 = self.ct_39 if self.ct_39 else 0
            ct_40b = self.ct_40b if self.ct_40b else 0
            ct_42 = self.ct_42 if self.ct_42 else 0

            ct_27 = ct_29 + ct_30 + ct_32 + ct_32a
            ct_28 = ct_31 + ct_33
            ct_34 = ct_26 + ct_27
            ct_35 = ct_28
            ct_36 = ct_35 - ct_25
            if (ct_36 - ct_22 + ct_37 - ct_38 - ct_39) > 0:
                ct_40a = ct_36 - ct_22 + ct_37 - ct_38 - ct_39
            if (ct_36 - ct_22 + ct_37 - ct_38 - ct_39) < 0:
                ct_41 = ct_36 - ct_22 + ct_37 - ct_38 - ct_39

        content = content.replace('var_kyKKhai', str(self.kyKKhai), 1)
        content = content.replace('var_kyKKhaiTuNgay', str(self.kyKKhaiTuNgay.strftime("%d/%m/%Y")), 1)
        content = content.replace('var_kyKKhaiDenNgay', str(self.kyKKhaiDenNgay.strftime("%d/%m/%Y")), 1)
        content = content.replace('var_ngayLapTKhai', str(date.today().strftime("%d/%m/%Y")), 1)
        content = content.replace('var_maCQTNoiNop', str(self.maCQTNoiNop), 1)
        content = content.replace('var_tenCQTNoiNop', str(self.tenCQTNoiNop), 1)
        content = content.replace('var_nguoiKy', str(self.nguoiKy), 1)
        content = content.replace('var_ngayKy', str(self.ngayKy.strftime("%d/%m/%Y")), 1)
        content = content.replace('var_nganhNgheKD', str(self.nganhNgheKD), 1)
        content = content.replace('var_ct21', str(ct_21))
        content = content.replace('var_ct22', str(ct_22))
        content = content.replace('var_ct23', str(ct_23))
        content = content.replace('var_ct24', str(ct_24))
        content = content.replace('var_ct25', str(ct_25))
        content = content.replace('var_ct26', str(ct_26))
        content = content.replace('var_ct27', str(ct_27))
        content = content.replace('var_ct28', str(ct_28))
        content = content.replace('var_ct29', str(ct_29))
        content = content.replace('var_ct30', str(ct_30))
        content = content.replace('var_ct31', str(ct_31))
        content = content.replace('var_ct32a', str(ct_32a))
        content = content.replace('var_ct32', str(ct_32))
        content = content.replace('var_ct33', str(ct_33))
        content = content.replace('var_ct34', str(ct_34))
        content = content.replace('var_ct35', str(ct_35))
        content = content.replace('var_ct36', str(abs(ct_36)))
        content = content.replace('var_ct37', str(ct_37))
        content = content.replace('var_ct38', str(ct_38))
        content = content.replace('var_ct39', str(ct_39))
        content = content.replace('var_ct40a', str(ct_40a), 1)
        content = content.replace('var_ct41', str(abs(ct_41)))
        content = content.replace('var_ct42', str(ct_42))
        content = content.replace('var_ct22', str(ct_22) if self.ct_22 else '0')
        content = content.replace('var_ct25', str(ct_25) if self.ct_25 else '0')
        content = content.replace('var_ct37', str(ct_37) if self.ct_37 else '0')
        content = content.replace('var_ct38', str(ct_38) if self.ct_38 else '0')
        content = content.replace('var_ct39', str(ct_39) if self.ct_39 else '0')
        content = content.replace('var_ct40b', str(ct_40b) if self.ct_40b else '0', 1)
        content = content.replace('var_ct42', str(ct_42) if self.ct_42 else '0')
        ct_40 = ct_40a - ct_40b
        if ct_41 < 0:
            ct_41 = abs(ct_41)
        else:
            ct_41 = 0
        ct_43 = ct_41 - ct_42
        if ct_43 < 0:
            raise UserError(
                "Tổng số thuế GTGT đề nghị hoàn [42] đang lớn thuế GTGT chưa khấu trừ hết kỳ này [41], yêu cầu kiểm tra lại.")
        else:
            self.ct_43 = ct_43
        content = content.replace('var_ct40', str(ct_40))
        content = content.replace('var_ct43', str(abs(ct_43)))

        information_tax = self.env['information.tax'].sudo().search([], limit=1)
        if information_tax:
            content = content.replace('var_mst', str(information_tax.mst) if information_tax.mst else '')
            content = content.replace('var_tenNNT', str(information_tax.tenNNT) if information_tax.tenNNT else '')
            content = content.replace('var_dchiNNT', str(information_tax.dchiNNT) if information_tax.dchiNNT else '')
            content = content.replace('var_tenHuyenNNT', str(information_tax.tenHuyenNNT) if information_tax.tenHuyenNNT else '')
            content = content.replace('var_tenTinhNNT', str(information_tax.tenTinhNNT) if information_tax.tenTinhNNT else '')
            content = content.replace('var_dthoaiNNT', str(information_tax.phone) if information_tax.phone else '')
            content = content.replace('var_faxNNT', str(information_tax.fax) if information_tax.fax else '')
            content = content.replace('var_emailNNT', str(information_tax.email) if information_tax.email else '')
            content = content.replace('var_ct11_ct', str(information_tax.name_dl_thue) if information_tax.name_dl_thue else '')
            content = content.replace('var_ct11_cct', str(information_tax.name_dl_thue) if information_tax.name_dl_thue else '')

        file_content.write(content)
        file_content.close()
        template_path = os.path.realpath(file_content.name)
        file_content = open(template_path, 'rb')
        data = base64.b64encode(file_content.read())
        #
        os.remove("TỜ KHAI THUẾ GIÁ TRỊ GIA TĂNG.xml")
        #
        # self.env['tax.report.attachment'].sudo().search([]).unlink()
        # file_output = self.env['tax.report.attachment'].sudo().create({
        #     'file': data,
        #     'file_name': "Tờ khai thuế kỳ " + str(self.kyKKhai) + ".xml"
        # })
        self.file = data
        self.file_name = "Tờ khai thuế kỳ " + str(self.kyKKhai) + ".xml"
        return {
            'type': 'ir.actions.act_url',
            'target': 'new',
            'url': 'web/content/?model=' + self._name + '&id=' + str(
                self.id) + '&field=file&download=true&filename=' + str(self.file_name),
        }

    def export_xlsx(self):
        options = {'unfolded_lines': [],
                   'date': {'string': '2020', 'period_type': 'month', 'mode': 'range',
                            'date_from': str(self.kyKKhaiTuNgay.strftime("%Y-%m-%d")),
                            'date_to': str(self.kyKKhaiDenNgay.strftime("%Y-%m-%d")), 'filter': 'this_month',
                            'strict_range': True},
                   'comparison': {'filter': 'no_comparison', 'number_period': 1, 'date_from': '', 'date_to': '',
                                  'periods': []}, 'all_entries': False, 'unposted_in_period': True}
        template_path = get_module_resource('advanced_vn_report', 'static/src/xml/template',
                                            'Báo cáo Thuế.xlsx')
        workbook = openpyxl.load_workbook(template_path)
        sheet = workbook['Sheet1']
        lines = self._get_lines(options)
        dict_sale = {}
        dict_purchase = {}
        # try:
        for line in lines:
            if '(10.0)' in line['name'] and self.env['account.tax'].sudo().search(
                    [('id', '=', line['id'])]).type_tax_use == 'sale':
                dict_sale['(10.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
            elif '(5.0)' in line['name'] and self.env['account.tax'].sudo().search(
                    [('id', '=', line['id'])]).type_tax_use == 'sale':
                dict_sale['(5.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
            elif '(0.0)' in line['name'] and self.env['account.tax'].sudo().search(
                    [('id', '=', line['id'])]).type_tax_use == 'sale':
                dict_sale['(0.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
            elif '(10.0)' in line['name'] and self.env['account.tax'].sudo().search(
                    [('id', '=', line['id'])]).type_tax_use == 'purchase':
                dict_purchase['(10.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
            elif '(5.0)' in line['name'] and self.env['account.tax'].sudo().search(
                    [('id', '=', line['id'])]).type_tax_use == 'purchase':
                dict_purchase['(5.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
            elif '(0.0)' in line['name'] and self.env['account.tax'].sudo().search(
                    [('id', '=', line['id'])]).type_tax_use == 'purchase':
                dict_purchase['(0.0)'] = [line['columns'][0]['name'], line['columns'][1]['name']]
        invoices = self.env['account.move'].sudo().search(
            [('state', '=', 'posted'), ('type', '=', 'out_invoice'), ('date', '<=', self.kyKKhaiDenNgay),
             ('date', '>=', self.kyKKhaiTuNgay)])
        out_refunds = self.env['account.move'].sudo().search(
            [('state', '=', 'posted'), ('type', '=', 'out_refund'), ('date', '<=', self.kyKKhaiDenNgay),
             ('date', '>=', self.kyKKhaiTuNgay)])
        non_tax_total = 0
        for invoice in invoices:
            if invoice.invoice_line_ids:
                for inline in invoice.invoice_line_ids:
                    if not inline.tax_ids:
                        non_tax_total += inline.price_subtotal
        for out_refund in out_refunds:
            if out_refund.invoice_line_ids:
                for inline in out_refund.invoice_line_ids:
                    if not inline.tax_ids:
                        non_tax_total -= inline.price_subtotal
        invoices_25 = self.env['account.move'].sudo().search(
            [('state', '=', 'posted'), ('type', '=', 'in_invoice'), ('nontaxable_income_source', '=', False),
             ('date', '<=', self.kyKKhaiDenNgay),
             ('date', '>=', self.kyKKhaiTuNgay)])
        invoices_25_refund = self.env['account.move'].sudo().search(
            [('state', '=', 'posted'), ('type', '=', 'in_refund'), ('nontaxable_income_source', '=', False),
             ('date', '<=', self.kyKKhaiDenNgay),
             ('date', '>=', self.kyKKhaiTuNgay)])
        tax_25 = 0
        for invoice_25 in invoices_25:
            tax_25 += invoice_25.amount_tax
        for invoice_25_re in invoices_25_refund:
            tax_25 -= invoice_25_re.amount_tax

        dau_phan_cach = self.env['res.lang'].sudo().search([('code', '=', self.env.user.lang)]).thousands_sep
        sheet['F31'] = float(
            dict_sale["(10.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '')) if '(10.0)' in dict_sale and \
                                                                                    dict_sale["(10.0)"] else 0
        sheet['H31'] = float(
            dict_sale["(10.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '')) if '(10.0)' in dict_sale and \
                                                                                    dict_sale["(10.0)"] else 0
        sheet['F30'] = float(
            dict_sale["(5.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '')) if '(5.0)' in dict_sale and \
                                                                                   dict_sale["(5.0)"] else 0
        sheet['H30'] = float(
            dict_sale["(5.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '')) if '(5.0)' in dict_sale and \
                                                                                   dict_sale["(5.0)"] else 0
        sheet['F29'] = float(
            dict_sale["(0.0)"][0].replace(' ₫', '').replace(dau_phan_cach, '')) if '(0.0)' in dict_sale and \
                                                                                   dict_sale["(0.0)"] else 0
        sheet['G29'] = float(
            dict_sale["(0.0)"][1].replace(' ₫', '').replace(dau_phan_cach, '')) if '(0.0)' in dict_sale and \
                                                                                   dict_sale["(0.0)"] else 0
        tong_von = 0
        in_invoices = self.env['account.move'].sudo().search(
            [('state', '=', 'posted'), ('type', '=', 'in_invoice'),
             ('date', '<=', self.kyKKhaiDenNgay),
             ('date', '>=', self.kyKKhaiTuNgay)])
        in_refunds = self.env['account.move'].sudo().search(
            [('state', '=', 'posted'), ('type', '=', 'in_refund'),
             ('date', '<=', self.kyKKhaiDenNgay),
             ('date', '>=', self.kyKKhaiTuNgay)])
        for in_refund in in_refunds:
            tong_von -= in_refund.amount_untaxed
        for in_invoice in in_invoices:
            tong_von += in_invoice.amount_untaxed
        tong_thue = 0
        for key in dict_purchase:
            # tong_von += float(dict_purchase[key][0].replace(' ₫', '').replace(dau_phan_cach, ''))
            tong_thue += float(dict_purchase[key][1].replace(' ₫', '').replace(dau_phan_cach, ''))
        sheet['F24'] = tong_von
        sheet['H24'] = tong_thue
        sheet['F27'] = non_tax_total
        sheet['H44'] = self.ct_42
        sheet['H21'] = self.ct_22
        # sheet['H25'] = self.ct_25
        sheet['H36'] = self.ct_37
        sheet['H37'] = self.ct_38
        sheet['H38'] = self.ct_39
        sheet['H44'] = self.ct_42
        sheet['H41'] = self.ct_40b
        dt_0_10 = sheet['F29'].value + sheet['F31'].value
        dt_khong_thue = sheet['F27'].value
        if (dt_0_10 + dt_khong_thue) > 0:
            sheet['H25'] = round(tax_25 / (dt_0_10 + dt_khong_thue) * dt_0_10)

        # try reload
        ct_41 = sheet['H30'].value + sheet['H31'].value - sheet['H25'].value - sheet['H21'].value + sheet['H36'].value - \
                sheet['H37'].value - sheet['H38'].value
        if ct_41 < 0:
            ct_41 = abs(ct_41)
        else:
            ct_41 = 0
        if (ct_41 - self.ct_42) < 0:
            raise UserError(
                "Tổng số thuế GTGT đề nghị hoàn [42] đang lớn thuế GTGT chưa khấu trừ hết kỳ này [41], yêu cầu kiểm tra lại.")
        else:
            self.ct_43 = ct_41 - self.ct_42

        # self.ct_43 = sheet['H45'].value
        # except Exception as e:
        #     raise UserError("Không thể xuất báo cáo.")
        docx_stream = io.BytesIO()
        workbook.save(docx_stream)

        out = base64.encodestring(docx_stream.getvalue())
        file_output = self.env['contract.file'].sudo().create({
            'file': out,
            'file_name': "Báo cáo Thuế" + ".xlsx",
            'file_type': 'thiet_ke'
        })
        # wb2 = load_workbook(file_output.file_name, data_only=True)
        # self.ct_43 = wb2['Sheet1']['H45'].value
        # docx_bytes = docx_stream.getvalue()
        return {
            'type': 'ir.actions.act_url',
            'target': 'new',
            'url': 'web/content/?model=' + file_output._name + '&id=' + str(
                file_output.id) + '&field=file&download=true&filename=' + str(file_output.file_name),
        }
